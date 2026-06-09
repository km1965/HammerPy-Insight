#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
HammerPy Insight - Script Principal (Phase 3)
---------------------------------------------
Application d'automatisation et d'interprétation des résultats transitoires
exportés depuis le logiciel d'hydraulique Bentley HAMMER.

Fonctionnalités :
  - Parsing ultra-robuste des fichiers CSV/Excel (formats HAMMER francophones et anglophones)
  - Interface moderne avec configuration de l'étude (Projet, Ingénieur, PN conduite)
  - Visualisation graphique interactive avec seuils critiques (Matplotlib)
  - Export de rapport professionnel Word (.docx) avec tableaux et graphique intégré

Auteur  : HammerPy Insight / Expert Python & CustomTkinter
Version : 3.0 - Juin 2026
"""

import os
import json
import tempfile
import io
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox

import pandas as pd
import customtkinter as ctk

# Imports pour Matplotlib et son intégration dans Tkinter
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

# ── Imports des modules extraits ───────────────────────────────────
from utils import (
    PN_CLASSES, PMIN_OPTIONS, FLOW_UNITS, VOLUME_UNITS,
    VOLUME_THRESHOLD_L_DEFAULT, FLOW_UNIT_HINTS, VOLUME_UNIT_HINTS,
)
from data_parser import HammerDataParser
from workbook import WorkbookManager
from pump_parser import PumpReportParser
from report_generator import WordReportGenerator, DOCX_AVAILABLE
from air_valve_sizing import AirValveSizing
from dxf_profile_importer import load_dxf_both, HAS_EZDXF
from column_mapper import get_mapper as _get_column_mapper
from column_mapper_dialog import ask_column_mapping as _ask_column_mapping
from ventouses_report import export_ventouses_report, DOCX_AVAILABLE as VENTOUSES_DOCX_AVAILABLE
from system_diagnostics import SystemDiagnostics, ICON as DIAG_ICON, CAT_A as DIAG_CAT_A

# ── Alias rétrocompatibilité ──────────────────────────────────────
from utils import parse_number as _parse_number
from utils import find_col_in_df as _find_col_in_df

# Configuration du thème visuel global (doit être en amont de toute création de widget)
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")


class HammerPyApp(ctk.CTk):
    """
    Fenêtre principale de l'application de bureau HammerPy Insight.
    Structure moderne basée sur CustomTkinter avec Sidebar de navigation
    et Tabview central.
    """

    def __init__(self):
        super().__init__()

        # ── Propriétés de la fenêtre ────────────────────────────────────
        self.title("HammerPy Insight v3 — Analyse Transitoire Hydraulique")
        self.geometry("1200x750")
        self.minsize(1000, 650)

        # ── Icône de l'application ─────────────────────────────────────
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hammerpy_icon.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)

        # ── Initialisation du parser ────────────────────────────────────
        self.parser = HammerDataParser()
        self.workbook_manager = WorkbookManager()

        # ── Multi-pompe : liste de PumpReportParser ───────────────────
        self.pump_parsers: list[PumpReportParser] = []
        self.pump_selected_index: int = -1  # index de la pompe sélectionnée
        self.var_pump_mode = tk.StringVar(value="Continu")

        # ── Phase 3 : Ventouses & Vidanges ──────────────────────────
        self.air_valve_sizer = AirValveSizing()
        self.var_pipe_dn = tk.StringVar(value="250")
        self.valve_profile_filepath: str = ""

        # ── Column Mapper (auto-apprentissage mapping colonnes) ─────
        self._column_mapper = _get_column_mapper()
        self._column_mapper.set_ui_callback(self._on_column_mapping_request)
        self.air_valve_sizer.set_column_mapper(self._column_mapper)

        # ── État de l'application ───────────────────────────────────────
        self.station_filepath: str = ""
        self.hpt_filepath: str = ""
        self.workbook_filepath: str = ""
        self.steady_state_status: dict | None = None
        self.transient_status: dict | None = None

        # ── Phase 4 : SystemDiagnostics (résultats du dernier run) ─────
        self.last_diagnostics: list[dict] | None = None
        self.last_diagnostics_summary: dict | None = None

        # ── Gestion du projet (fichier .hpi) ───────────────────────────
        self.current_project_path: str | None = None
        self.is_dirty: bool = False
        self.project_creation_date: str = datetime.now().isoformat(timespec="seconds")
        self._suppress_dirty: bool = False  # Anti-boucle lors du chargement

        # ── Unités d'affichage (préférence UI, pas de "dirty") ────────
        self.var_flow_unit       = tk.StringVar(value="m³/h")
        self.var_volume_unit     = tk.StringVar(value="L")
        self.var_volume_threshold = tk.StringVar(value=str(VOLUME_THRESHOLD_L_DEFAULT))
        # Source units (override manuel ; "Auto" = auto-détection)
        self.var_station_src_unit   = tk.StringVar(value="Auto-détection")
        self.var_transient_src_unit = tk.StringVar(value="Auto-détection")

        # ── Objets Matplotlib ───────────────────────────────────────────
        self.canvas = None
        self.toolbar = None
        self.current_fig = None  # Référence à la figure active pour l'export

        # ── Layout de la fenêtre principale ────────────────────────────
        self.grid_columnconfigure(0, weight=0)   # Sidebar fixe
        self.grid_columnconfigure(1, weight=1)   # Zone principale extensible
        self.grid_rowconfigure(0, weight=0)      # Barre de menu supérieure (hauteur fixe)
        self.grid_rowconfigure(1, weight=1)      # Zone principale extensible

        # ── Construction des composants ─────────────────────────────────
        self._create_top_bar()
        self._create_sidebar()
        self._create_main_content()
        self._update_title()

    # ================================================================
    # BARRE DE MENU SUPÉRIEURE (Ouvrir / Enregistrer / Enregistrer sous / Quitter)
    # ================================================================

    def _create_top_bar(self):
        """Crée la barre horizontale de boutons de gestion de projet."""
        self.top_bar = ctk.CTkFrame(self, height=50, corner_radius=0, fg_color=("gray90", "gray17"))
        self.top_bar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=0, pady=0)
        self.top_bar.grid_propagate(False)
        self.top_bar.grid_columnconfigure(4, weight=1)  # Espace central extensible

        # ── Bouton Ouvrir ──────────────────────────────────────────────
        self.btn_open = ctk.CTkButton(
            self.top_bar,
            text="📂  Ouvrir",
            width=130,
            fg_color="#1f538d", hover_color="#14375e",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._open_project,
        )
        self.btn_open.grid(row=0, column=0, padx=(14, 6), pady=8, sticky="w")

        # ── Bouton Enregistrer ────────────────────────────────────────
        self.btn_save = ctk.CTkButton(
            self.top_bar,
            text="💾  Enregistrer",
            width=140,
            fg_color="#2d7d3a", hover_color="#1f5a28",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._save_project_quick,
        )
        self.btn_save.grid(row=0, column=1, padx=6, pady=8, sticky="w")

        # ── Bouton Enregistrer sous ───────────────────────────────────
        self.btn_save_as = ctk.CTkButton(
            self.top_bar,
            text="💾  Enregistrer sous…",
            width=160,
            fg_color="transparent", border_width=1,
            text_color=("gray10", "gray90"),
            hover_color=("gray75", "gray30"),
            font=ctk.CTkFont(size=12),
            command=self._save_project_as,
        )
        self.btn_save_as.grid(row=0, column=2, padx=6, pady=8, sticky="w")

        # ── Séparateur visuel ─────────────────────────────────────────
        sep = ctk.CTkFrame(self.top_bar, width=2, height=30, fg_color="gray40")
        sep.grid(row=0, column=3, padx=14, pady=10, sticky="w")

        # ── Titre central ─────────────────────────────────────────────
        self.lbl_top_title = ctk.CTkLabel(
            self.top_bar,
            text="⚡ HammerPy Insight  v3.0",
            font=ctk.CTkFont(family="Consolas", size=13, weight="bold"),
            text_color="gray",
        )
        self.lbl_top_title.grid(row=0, column=4, padx=10, pady=8, sticky="w")

        # ── Indicateur d'état + chemin projet (colonne extensible droite)
        self.top_bar.grid_columnconfigure(5, weight=1)
        self.lbl_project_state = ctk.CTkLabel(
            self.top_bar,
            text="● Nouveau projet (non sauvegardé)",
            font=ctk.CTkFont(size=11),
            text_color="orange",
            anchor="e",
        )
        self.lbl_project_state.grid(row=0, column=5, padx=10, pady=8, sticky="e")

        # ── Bouton Quitter ────────────────────────────────────────────
        self.btn_quit = ctk.CTkButton(
            self.top_bar,
            text="🚪  Quitter",
            width=110,
            fg_color="#a52828", hover_color="#7a1d1d",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._quit_app,
        )
        self.btn_quit.grid(row=0, column=6, padx=(6, 14), pady=8, sticky="e")

    # ================================================================
    # SIDEBAR
    # ================================================================

    def _create_sidebar(self):
        """Crée la barre latérale gauche de navigation."""
        self.sidebar_frame = ctk.CTkFrame(self, width=240, corner_radius=0)
        self.sidebar_frame.grid(row=1, column=0, sticky="nsew")
        self.sidebar_frame.grid_propagate(False)
        self.sidebar_frame.grid_rowconfigure(15, weight=1)  # Espace élastique en bas

        # Logo
        ctk.CTkLabel(
            self.sidebar_frame,
            text="⚡ HammerPy\nInsight",
            font=ctk.CTkFont(family="Consolas", size=22, weight="bold"),
            justify="center"
        ).grid(row=0, column=0, padx=20, pady=(30, 6))

        ctk.CTkLabel(
            self.sidebar_frame,
            text="v3.0 — Bentley HAMMER",
            font=ctk.CTkFont(size=10),
            text_color="gray"
        ).grid(row=1, column=0, padx=20, pady=(0, 24))

        # Séparateur
        ctk.CTkFrame(self.sidebar_frame, height=1, fg_color="gray40").grid(
            row=2, column=0, padx=15, sticky="ew")

        # Section Paramètres de la conduite
        ctk.CTkLabel(
            self.sidebar_frame,
            text="CONDUITE",
            font=ctk.CTkFont(size=10, weight="bold"),
            text_color="gray"
        ).grid(row=3, column=0, padx=20, pady=(18, 4), sticky="w")

        # Classe PN
        ctk.CTkLabel(self.sidebar_frame, text="Classe PN :", anchor="w").grid(
            row=4, column=0, padx=20, pady=(4, 0), sticky="w")
        self.var_pn = tk.StringVar(value="PN 16")
        self.opt_pn = ctk.CTkOptionMenu(
            self.sidebar_frame,
            values=list(PN_CLASSES.keys()),
            variable=self.var_pn,
            command=self._on_pn_change
        )
        self.opt_pn.grid(row=5, column=0, padx=20, pady=(2, 8), sticky="ew")

        # Pression min admissible
        ctk.CTkLabel(self.sidebar_frame, text="P min admissible :", anchor="w").grid(
            row=6, column=0, padx=20, pady=(4, 0), sticky="w")
        self.var_pmin = tk.StringVar(value=list(PMIN_OPTIONS.keys())[1])
        self.opt_pmin = ctk.CTkOptionMenu(
            self.sidebar_frame,
            values=list(PMIN_OPTIONS.keys()),
            variable=self.var_pmin,
            command=self._on_pmin_change,
            width=200
        )
        self.opt_pmin.grid(row=7, column=0, padx=20, pady=(2, 8), sticky="new")

        # Séparateur
        ctk.CTkFrame(self.sidebar_frame, height=1, fg_color="gray40").grid(
            row=8, column=0, padx=15, pady=8, sticky="ew")

        # Thème
        ctk.CTkLabel(
            self.sidebar_frame,
            text="AFFICHAGE",
            font=ctk.CTkFont(size=10, weight="bold"),
            text_color="gray"
        ).grid(row=9, column=0, padx=20, pady=(4, 4), sticky="w")

        ctk.CTkLabel(self.sidebar_frame, text="Mode :", anchor="w").grid(
            row=10, column=0, padx=20, pady=(0, 0), sticky="w")
        self.theme_opt = ctk.CTkOptionMenu(
            self.sidebar_frame,
            values=["Dark", "Light", "System"],
            command=self._change_appearance_mode
        )
        self.theme_opt.grid(row=11, column=0, padx=20, pady=(2, 12), sticky="ew")

        # Séparateur
        ctk.CTkFrame(self.sidebar_frame, height=1, fg_color="gray40").grid(
            row=12, column=0, padx=15, pady=4, sticky="ew")

        # Section Unités
        ctk.CTkLabel(
            self.sidebar_frame,
            text="UNITÉS D'AFFICHAGE",
            font=ctk.CTkFont(size=10, weight="bold"),
            text_color="gray"
        ).grid(row=13, column=0, padx=20, pady=(8, 4), sticky="w")

        ctk.CTkLabel(self.sidebar_frame, text="Débit :", anchor="w").grid(
            row=14, column=0, padx=20, pady=(2, 0), sticky="w")
        self.opt_flow_unit = ctk.CTkOptionMenu(
            self.sidebar_frame,
            values=list(FLOW_UNITS.keys()),
            variable=self.var_flow_unit,
            command=self._on_flow_unit_change,
            width=200
        )
        self.opt_flow_unit.grid(row=15, column=0, padx=20, pady=(2, 6), sticky="new")

        ctk.CTkLabel(self.sidebar_frame, text="Volume :", anchor="w").grid(
            row=16, column=0, padx=20, pady=(4, 0), sticky="w")
        self.opt_volume_unit = ctk.CTkOptionMenu(
            self.sidebar_frame,
            values=list(VOLUME_UNITS.keys()),
            variable=self.var_volume_unit,
            command=self._on_volume_unit_change,
            width=200
        )
        self.opt_volume_unit.grid(row=17, column=0, padx=20, pady=(2, 6), sticky="new")

        ctk.CTkLabel(self.sidebar_frame, text="Seuil HPT :", anchor="w").grid(
            row=18, column=0, padx=20, pady=(4, 0), sticky="w")
        self.entry_volume_threshold = ctk.CTkEntry(
            self.sidebar_frame,
            textvariable=self.var_volume_threshold,
            width=200
        )
        self.entry_volume_threshold.grid(row=19, column=0, padx=20, pady=(2, 4), sticky="new")
        self.lbl_threshold_unit = ctk.CTkLabel(
            self.sidebar_frame,
            text="(en L)",
            font=ctk.CTkFont(size=10),
            text_color="gray"
        )
        self.lbl_threshold_unit.grid(row=20, column=0, padx=20, pady=(0, 6), sticky="new")

        # Hooks de modification
        self.var_volume_threshold.trace_add("write",
            lambda *_: self._on_threshold_change())

        # Bouton aide
        ctk.CTkButton(
            self.sidebar_frame,
            text="? Aide",
            fg_color="transparent",
            border_width=1,
            text_color=("gray20", "gray80"),
            hover_color=("gray75", "gray30"),
            command=self._show_help
        ).grid(row=21, column=0, padx=20, pady=(4, 20), sticky="s")

    # ================================================================
    # TABVIEW PRINCIPAL
    # ================================================================

    def _create_main_content(self):
        """Crée la zone centrale avec les onglets."""
        self.tabview = ctk.CTkTabview(self, corner_radius=10)
        self.tabview.grid(row=1, column=1, padx=(10, 18), pady=15, sticky="nsew")

        for name in ["Régime Permanent", "Analyse Transitoire", "Rapport Technique",
                       "Ventouses & Vidanges", "Système & Diagnostics"]:
            self.tabview.add(name)

        self.tabview._segmented_button.configure(font=ctk.CTkFont(size=13, weight="bold"))

        self._setup_steady_state_tab()
        self._setup_transient_tab()
        self._setup_report_tab()
        self._setup_valve_tab()
        self._setup_diagnostics_tab()

        # ── Hooks de détection de modifications utilisateur ──────────
        # Entry du projet / ingénieur → dirty à chaque frappe
        self.entry_projet.bind("<KeyRelease>",     lambda _e: self._mark_dirty())
        self.entry_projet.bind("<FocusOut>",       lambda _e: self._mark_dirty())
        self.entry_ingenieur.bind("<KeyRelease>",  lambda _e: self._mark_dirty())
        self.entry_ingenieur.bind("<FocusOut>",    lambda _e: self._mark_dirty())

        # Textbox du rapport → dirty à chaque édition
        def _on_report_modified(_event=None):
            if self.txt_report.edit_modified():
                self._mark_dirty()
                self.txt_report.edit_modified(False)
        self.txt_report.bind("<<Modified>>", _on_report_modified)

    # ──────────────────────────────────────────────────────────────────
    # Onglet 1 : Régime Permanent
    # ──────────────────────────────────────────────────────────────────
    def _setup_steady_state_tab(self):
        tab = self.tabview.tab("Régime Permanent")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(tab, text="Caractérisation du Régime Permanent",
                     font=ctk.CTkFont(size=18, weight="bold"), anchor="w"
                     ).grid(row=0, column=0, padx=20, pady=(18, 4), sticky="w")

        # ── Panneau de configuration du projet ───────────────────────
        cfg_frame = ctk.CTkFrame(tab)
        cfg_frame.grid(row=1, column=0, padx=20, pady=8, sticky="ew")
        cfg_frame.grid_columnconfigure((1, 3), weight=1)

        ctk.CTkLabel(cfg_frame, text="Nom du projet :").grid(
            row=0, column=0, padx=(15, 6), pady=12, sticky="w")
        self.entry_projet = ctk.CTkEntry(cfg_frame, placeholder_text="Ex : AEP Ville-Nord — Adduction principale")
        self.entry_projet.grid(row=0, column=1, padx=(0, 20), pady=12, sticky="ew")

        ctk.CTkLabel(cfg_frame, text="Ingénieur :").grid(
            row=0, column=2, padx=(0, 6), pady=12, sticky="w")
        self.entry_ingenieur = ctk.CTkEntry(cfg_frame, placeholder_text="Nom de l'ingénieur")
        self.entry_ingenieur.grid(row=0, column=3, padx=(0, 15), pady=12, sticky="ew")

        # ── Import du fichier station ─────────────────────────────────
        import_frame = ctk.CTkFrame(tab)
        import_frame.grid(row=2, column=0, padx=20, pady=8, sticky="nsew")
        import_frame.grid_columnconfigure(1, weight=1)
        import_frame.grid_rowconfigure(3, weight=1)

        ctk.CTkLabel(
            import_frame,
            text="Chargez le fichier de la station pour extraire le débit nominal (Q) et la HMT.",
            text_color="gray", anchor="w"
        ).grid(row=0, column=0, columnspan=3, padx=18, pady=(16, 6), sticky="w")

        self.btn_import_steady = ctk.CTkButton(
            import_frame, text="  Charger fichier station  (.csv / .xlsx)",
            command=self._import_steady_state_file
        )
        self.btn_import_steady.grid(row=1, column=0, padx=18, pady=8, sticky="w")

        self.lbl_steady_file = ctk.CTkLabel(import_frame, text="Aucun fichier sélectionné",
                                            text_color="gray", anchor="w")
        self.lbl_steady_file.grid(row=1, column=1, padx=8, pady=8, sticky="ew")

        # Sélecteur d'unité source pour le débit
        ctk.CTkLabel(import_frame, text="Unité :", anchor="w",
                     text_color="gray").grid(row=1, column=2, padx=(0, 4), pady=8, sticky="e")
        self.opt_station_src_unit = ctk.CTkOptionMenu(
            import_frame,
            values=["Auto-détection", "m³/h", "L/s"],
            variable=self.var_station_src_unit,
            width=150
        )
        self.opt_station_src_unit.grid(row=1, column=3, padx=(0, 18), pady=8, sticky="e")

        # ── Carte de résultats ────────────────────────────────────────
        kpi_frame = ctk.CTkFrame(import_frame, fg_color=("gray88", "gray14"))
        kpi_frame.grid(row=2, column=0, columnspan=3, padx=18, pady=16, sticky="ew")
        kpi_frame.grid_columnconfigure((0, 1), weight=1)

        for col, title, attr in [(0, "Débit Nominal (Q)", "lbl_flow_val"),
                                  (1, "Hauteur Manométrique Totale (HMT)", "lbl_hmt_val")]:
            ctk.CTkLabel(kpi_frame, text=title, font=ctk.CTkFont(size=12, weight="bold")
                         ).grid(row=0, column=col, padx=24, pady=(18, 2))
            lbl = ctk.CTkLabel(kpi_frame, text="-- --",
                               font=ctk.CTkFont(size=28, weight="bold"), text_color="#1f8ecf")
            lbl.grid(row=1, column=col, padx=24, pady=(2, 18))
            setattr(self, attr, lbl)

        # Zone de détails (colonnes du fichier)
        self.lbl_steady_details = ctk.CTkLabel(
            import_frame, text="", text_color="gray", anchor="w", justify="left")
        self.lbl_steady_details.grid(row=3, column=0, columnspan=3, padx=18, pady=(4, 12), sticky="ew")

    # ──────────────────────────────────────────────────────────────────
    # Onglet 2 : Analyse Transitoire
    # ──────────────────────────────────────────────────────────────────
    def _setup_transient_tab(self):
        tab = self.tabview.tab("Analyse Transitoire")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(6, weight=1)

        ctk.CTkLabel(tab, text="Analyse des Pressions & Volumes Transitoires",
                     font=ctk.CTkFont(size=18, weight="bold"), anchor="w"
                     ).grid(row=0, column=0, padx=20, pady=(18, 4), sticky="w")

        # ── Section 1 : Import HPT (fichier CSV/Excel existant) ──────
        frame = ctk.CTkFrame(tab)
        frame.grid(row=1, column=0, padx=20, pady=8, sticky="ew")
        frame.grid_columnconfigure(1, weight=1)

        self.btn_import_transient = ctk.CTkButton(
            frame, text="  Importer données HPT  (.csv / .xlsx)",
            fg_color="#1f538d", hover_color="#14375e",
            command=self._import_transient_file
        )
        self.btn_import_transient.grid(row=0, column=0, padx=14, pady=12, sticky="w")

        self.lbl_transient_file = ctk.CTkLabel(frame, text="Aucun fichier HPT chargé",
                                               text_color="gray", anchor="w")
        self.lbl_transient_file.grid(row=0, column=1, padx=8, pady=12, sticky="ew")

        ctk.CTkLabel(frame, text="Unité :", anchor="e",
                     text_color="gray").grid(row=0, column=2, padx=(0, 4), pady=12, sticky="e")
        self.opt_transient_src_unit = ctk.CTkOptionMenu(
            frame,
            values=["Auto-détection", "L", "m³"],
            variable=self.var_transient_src_unit,
            width=150
        )
        self.opt_transient_src_unit.grid(row=0, column=3, padx=(0, 14), pady=12, sticky="e")

        # ── KPI strip (pressions + volume) ───────────────────────────
        kpi_strip = ctk.CTkFrame(tab, fg_color=("gray88", "gray14"))
        kpi_strip.grid(row=2, column=0, padx=20, pady=(0, 8), sticky="ew")
        kpi_strip.grid_rowconfigure(0, weight=0)
        kpi_strip.grid_columnconfigure((0, 1, 2), weight=1)

        for col, title, attr, unit in [
            (0, "Pression Min (Dépression)", "lbl_pmin_val", "bar"),
            (1, "Pression Max (Surpression)", "lbl_pmax_val", "bar"),
            (2, "Volume Gaz Max (HPT)", "lbl_vgas_val", "L"),
        ]:
            ctk.CTkLabel(kpi_strip, text=title, font=ctk.CTkFont(size=11)
                         ).grid(row=0, column=col, pady=(12, 0))
            lbl = ctk.CTkLabel(kpi_strip, text=f"-- {unit}",
                               font=ctk.CTkFont(size=20, weight="bold"), text_color="#1f8ecf")
            lbl.grid(row=1, column=col, pady=(0, 12))
            setattr(self, attr, lbl)

        # ── Section 2 : Classeur HAMMER (Flex Tables .xlsx/.xls) ─────
        wb_frame = ctk.CTkFrame(tab, fg_color=("gray88", "gray14"))
        wb_frame.grid(row=3, column=0, padx=20, pady=8, sticky="ew")
        wb_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(wb_frame, text="MODÈLE HAMMER",
                     font=ctk.CTkFont(size=11, weight="bold"), text_color="gray"
                     ).grid(row=0, column=0, columnspan=4, padx=18, pady=(12, 4), sticky="w")

        self.btn_import_workbook = ctk.CTkButton(
            wb_frame, text="  Charger classeur HAMMER  (.xlsx / .xls)",
            fg_color="#2d7d3a", hover_color="#1f5a28",
            command=self._import_workbook
        )
        self.btn_import_workbook.grid(row=1, column=0, padx=14, pady=8, sticky="w")

        self.lbl_workbook_file = ctk.CTkLabel(wb_frame, text="Aucun classeur chargé",
                                              text_color="gray", anchor="w")
        self.lbl_workbook_file.grid(row=1, column=1, padx=8, pady=8, sticky="ew")

        self.btn_reset_workbook = ctk.CTkButton(
            wb_frame, text="Réinitialiser", width=100,
            fg_color="transparent", border_width=1,
            text_color=("gray10", "gray90"),
            hover_color=("gray75", "gray30"),
            state="disabled",
            command=self._reset_workbook
        )
        self.btn_reset_workbook.grid(row=1, column=2, padx=(0, 4), pady=8, sticky="e")

        # ── Compteurs de feuilles ────────────────────────────────────
        counter_frame = ctk.CTkFrame(wb_frame, fg_color="transparent")
        counter_frame.grid(row=2, column=0, columnspan=4, padx=14, pady=(0, 8), sticky="ew")
        counter_frame.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)

        self._wb_counter_labels = {}
        for col, (key, label) in enumerate([
            ("pipes", "Pipes"), ("nodes", "Nœuds"), ("pumps", "Pompes"),
            ("reservoirs", "Réservoirs"), ("hpt", "HPT"), ("air_valves", "Ventouses"),
        ]):
            ctk.CTkLabel(counter_frame, text=label, font=ctk.CTkFont(size=10),
                         text_color="gray").grid(row=0, column=col, pady=(4, 0))
            lbl = ctk.CTkLabel(counter_frame, text="—",
                               font=ctk.CTkFont(size=14, weight="bold"), text_color="#1f8ecf")
            lbl.grid(row=1, column=col, pady=(0, 4))
            self._wb_counter_labels[key] = lbl

        # ── Mini-stats du modèle ─────────────────────────────────────
        stats_frame = ctk.CTkFrame(wb_frame, fg_color="transparent")
        stats_frame.grid(row=3, column=0, columnspan=4, padx=14, pady=(0, 12), sticky="ew")
        stats_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self._wb_stats = {}
        for col, (key, label, unit) in enumerate([
            ("pmax", "P max", "bar"),
            ("pmin", "P min", "bar"),
            ("vmax_pump", "Q max pompe", "L/s"),
            ("materials", "Matériaux", ""),
        ]):
            ctk.CTkLabel(stats_frame, text=label, font=ctk.CTkFont(size=10),
                         text_color="gray").grid(row=0, column=col, pady=(4, 0))
            lbl = ctk.CTkLabel(stats_frame, text=f"— {unit}" if unit else "—",
                               font=ctk.CTkFont(size=12, weight="bold"), text_color="#1f8ecf")
            lbl.grid(row=1, column=col, pady=(0, 4))
            self._wb_stats[key] = lbl

        # ── Section 3 : BATTERIE POMPES — Mode + liste + courbe H(Q) ──
        pump_frame = ctk.CTkFrame(tab, fg_color=("gray88", "gray14"))
        pump_frame.grid(row=4, column=0, padx=20, pady=8, sticky="ew")
        pump_frame.grid_columnconfigure(1, weight=1)

        # ── Ligne 0 : titre + mode de fonctionnement ─────────────────
        ctk.CTkLabel(pump_frame, text="BATTERIE DE POMPES",
                     font=ctk.CTkFont(size=11, weight="bold"), text_color="gray"
                     ).grid(row=0, column=0, padx=18, pady=(12, 4), sticky="w")

        ctk.CTkLabel(pump_frame, text="Mode :", font=ctk.CTkFont(size=10),
                     text_color="gray").grid(row=0, column=2, padx=(0, 4), pady=(12, 4), sticky="e")
        self.opt_pump_mode = ctk.CTkOptionMenu(
            pump_frame, variable=self.var_pump_mode,
            values=["Continu", "Parallèle"], width=130,
            command=lambda _: self._on_pump_mode_change()
        )
        self.opt_pump_mode.grid(row=0, column=3, padx=(0, 14), pady=(12, 4), sticky="e")

        # ── Ligne 1 : boutons charger / supprimer + label fichier ────
        self.btn_import_pump = ctk.CTkButton(
            pump_frame, text="+ Charger rapport pompe (.rtf/.txt)",
            fg_color="#8b5cf6", hover_color="#6d28d9",
            command=self._import_pump_report
        )
        self.btn_import_pump.grid(row=1, column=0, padx=14, pady=8, sticky="w")

        self.btn_remove_pump = ctk.CTkButton(
            pump_frame, text="– Retirer la pompe sélectionnée", width=180,
            fg_color="#a52828", hover_color="#7a1d1d",
            state="disabled",
            command=self._remove_pump
        )
        self.btn_remove_pump.grid(row=1, column=1, padx=(0, 4), pady=8, sticky="w")

        self.btn_import_pump_data = ctk.CTkButton(
            pump_frame, text="📋 Importer données pompe (XLSX/CSV)", width=220,
            fg_color="#2e7d32", hover_color="#1b5e20",
            command=self._import_pump_data
        )
        self.btn_import_pump_data.grid(row=1, column=2, padx=(0, 4), pady=8, sticky="w")

        self.btn_gen_templates = ctk.CTkButton(
            pump_frame, text="📄 Générer templates XLSX", width=180,
            fg_color="#1565c0", hover_color="#0d47a1",
            command=self._generate_templates
        )
        self.btn_gen_templates.grid(row=1, column=3, padx=(0, 14), pady=8, sticky="w")

        # ── Ligne 2 : liste des pompes chargées ──────────────────────
        self.lst_pumps = ctk.CTkTextbox(
            pump_frame, height=70, font=ctk.CTkFont(family="Consolas", size=10),
            state="disabled", fg_color=("gray92", "gray18")
        )
        self.lst_pumps.grid(row=2, column=0, columnspan=4, padx=14, pady=(0, 8), sticky="ew")
        self.lst_pumps.bind("<ButtonRelease-1>", self._on_pump_list_click)

        # ── Ligne 3 : KPI strip pompe sélectionnée ───────────────────
        pump_kpi = ctk.CTkFrame(pump_frame, fg_color="transparent")
        pump_kpi.grid(row=3, column=0, columnspan=4, padx=14, pady=(0, 8), sticky="ew")
        pump_kpi.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)

        self._pump_kpi = {}
        for col, (key, label, unit) in enumerate([
            ("idx", "#", ""),
            ("label", "Pompe", ""),
            ("flow", "Q nom.", "L/s"),
            ("head", "HMT", "m"),
            ("npsh_a", "NPSH", "m"),
            ("n_pts", "Pts", ""),
        ]):
            ctk.CTkLabel(pump_kpi, text=label, font=ctk.CTkFont(size=10),
                         text_color="gray").grid(row=0, column=col, pady=(4, 0))
            lbl = ctk.CTkLabel(pump_kpi, text="—",
                               font=ctk.CTkFont(size=11, weight="bold"), text_color="#8b5cf6")
            lbl.grid(row=1, column=col, pady=(0, 4))
            self._pump_kpi[key] = lbl

        # ── Ligne 4 : saisie points de courbe H(Q) ──────────────────
        pts_label = ctk.CTkLabel(pump_frame, text="SAISIE POINTS DE COURBE (pompe sélectionnée)",
                                 font=ctk.CTkFont(size=10, weight="bold"), text_color="gray")
        pts_label.grid(row=4, column=0, columnspan=4, padx=18, pady=(8, 2), sticky="w")

        pts_row = ctk.CTkFrame(pump_frame, fg_color="transparent")
        pts_row.grid(row=5, column=0, columnspan=4, padx=14, pady=(0, 4), sticky="ew")
        pts_row.grid_columnconfigure(5, weight=1)

        ctk.CTkLabel(pts_row, text="Q (L/s) :", font=ctk.CTkFont(size=10),
                     text_color="gray").grid(row=0, column=0, padx=(0, 4), pady=4, sticky="e")
        self.entry_pump_q = ctk.CTkEntry(pts_row, width=80, placeholder_text="ex: 75")
        self.entry_pump_q.grid(row=0, column=1, padx=(0, 12), pady=4, sticky="w")

        ctk.CTkLabel(pts_row, text="H (m) :", font=ctk.CTkFont(size=10),
                     text_color="gray").grid(row=0, column=2, padx=(0, 4), pady=4, sticky="e")
        self.entry_pump_h = ctk.CTkEntry(pts_row, width=80, placeholder_text="ex: 150")
        self.entry_pump_h.grid(row=0, column=3, padx=(0, 12), pady=4, sticky="w")

        self.btn_add_pump_point = ctk.CTkButton(
            pts_row, text="+ Ajouter", width=90,
            fg_color="#8b5cf6", hover_color="#6d28d9",
            state="disabled",
            command=self._on_add_pump_point
        )
        self.btn_add_pump_point.grid(row=0, column=4, padx=(0, 6), pady=4, sticky="w")

        self.btn_clear_pump_points = ctk.CTkButton(
            pts_row, text="Effacer tout", width=90,
            fg_color="transparent", border_width=1,
            text_color=("gray10", "gray90"),
            hover_color=("gray75", "gray30"),
            state="disabled",
            command=self._on_clear_pump_points
        )
        self.btn_clear_pump_points.grid(row=0, column=5, padx=0, pady=4, sticky="w")

        # ── Ligne 6 : liste des points saisis ────────────────────────
        self.txt_pump_points = ctk.CTkTextbox(
            pump_frame, height=60, font=ctk.CTkFont(family="Consolas", size=10),
            state="disabled", fg_color=("gray92", "gray18")
        )
        self.txt_pump_points.grid(row=6, column=0, columnspan=4, padx=14, pady=(0, 8), sticky="ew")

        # ── Graphique H(Q) — toutes les pompes ──────────────────────
        self.pump_curve_frame = ctk.CTkFrame(tab, fg_color=("gray85", "gray12"))
        self.pump_curve_frame.grid(row=5, column=0, padx=20, pady=(0, 8), sticky="nsew")
        self.pump_curve_frame.grid_rowconfigure(0, weight=1)
        self.pump_curve_frame.grid_columnconfigure(0, weight=1)

        self.lbl_pump_curve_placeholder = ctk.CTkLabel(
            self.pump_curve_frame,
            text="📈   La courbe H(Q) s'affichera ici après chargement d'au moins 1 pompe avec ≥ 2 points.",
            text_color="gray", justify="center"
        )
        self.lbl_pump_curve_placeholder.grid(row=0, column=0, padx=20, pady=30, sticky="nsew")

        self.pump_curve_canvas = None
        self.pump_curve_toolbar = None

        # ── Zone graphique Matplotlib (HPT) ────────────────────────
        self.plot_placeholder_frame = ctk.CTkFrame(tab, fg_color=("gray85", "gray12"))
        self.plot_placeholder_frame.grid(row=6, column=0, padx=20, pady=(0, 12), sticky="nsew")
        self.plot_placeholder_frame.grid_rowconfigure(0, weight=1)
        self.plot_placeholder_frame.grid_columnconfigure(0, weight=1)

        self.lbl_graph_placeholder = ctk.CTkLabel(
            self.plot_placeholder_frame,
            text="📈   Le graphique interactif s'affichera ici après l'importation du fichier HPT.",
            text_color="gray", justify="center"
        )
        self.lbl_graph_placeholder.grid(row=0, column=0, padx=20, pady=40, sticky="nsew")

    # ──────────────────────────────────────────────────────────────────
    # Onglet 3 : Rapport Technique
    # ──────────────────────────────────────────────────────────────────
    def _setup_report_tab(self):
        tab = self.tabview.tab("Rapport Technique")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(tab, text="Rapport Technique — Prévisualisation & Export",
                     font=ctk.CTkFont(size=18, weight="bold"), anchor="w"
                     ).grid(row=0, column=0, padx=20, pady=(18, 4), sticky="w")

        frame = ctk.CTkFrame(tab)
        frame.grid(row=1, column=0, padx=20, pady=8, sticky="nsew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(
            frame,
            text="Prévisualisation du rapport.  Vous pouvez éditer le texte avant export.",
            text_color="gray", anchor="w"
        ).grid(row=0, column=0, padx=18, pady=(12, 4), sticky="w")

        self.txt_report = ctk.CTkTextbox(frame, font=ctk.CTkFont(family="Consolas", size=12), wrap="word")
        self.txt_report.grid(row=1, column=0, padx=18, pady=(4, 8), sticky="nsew")
        self._update_report_preview()

        # Barre de boutons d'export
        btn_bar = ctk.CTkFrame(frame, fg_color="transparent")
        btn_bar.grid(row=2, column=0, padx=18, pady=(4, 16), sticky="ew")
        btn_bar.grid_columnconfigure(0, weight=1)

        self.btn_export_txt = ctk.CTkButton(
            btn_bar, text="Export .txt",
            fg_color="transparent", border_width=1,
            text_color=("gray10", "gray90"),
            command=self._export_txt
        )
        self.btn_export_txt.grid(row=0, column=1, padx=6)

        self.btn_export_docx = ctk.CTkButton(
            btn_bar,
            text="  Générer Rapport Word (.docx)  ",
            fg_color="#1f538d", hover_color="#14375e",
            font=ctk.CTkFont(weight="bold"),
            state="normal" if DOCX_AVAILABLE else "disabled",
            command=self._export_word
        )
        self.btn_export_docx.grid(row=0, column=2, padx=6)

        if not DOCX_AVAILABLE:
            ctk.CTkLabel(btn_bar, text="python-docx non installé",
                         text_color="orange").grid(row=0, column=3, padx=6)

    # ------------------------------------------------------------------
    # Onglet 4 — Ventouses & Vidanges (Phase 3)
    # ------------------------------------------------------------------
    def _setup_valve_tab(self):
        tab = self.tabview.tab("Ventouses & Vidanges")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(2, weight=1)

        # ── Titre ───────────────────────────────────────────────────
        ctk.CTkLabel(tab, text="Profil en Long & Dimensionnement Ventouses + Vidanges",
                     font=ctk.CTkFont(size=18, weight="bold"), anchor="w"
                     ).grid(row=0, column=0, padx=20, pady=(18, 4), sticky="w")

        # ── Barre d'outils ─────────────────────────────────────────
        toolbar = ctk.CTkFrame(tab, fg_color="transparent")
        toolbar.grid(row=1, column=0, padx=20, pady=4, sticky="ew")
        toolbar.grid_columnconfigure(2, weight=1)

        ctk.CTkButton(toolbar, text="Importer profil (.csv)",
                      command=self._import_valve_profile
                      ).grid(row=0, column=0, padx=(0, 6))
        ctk.CTkButton(toolbar, text="CSV Bentley (FlexTable)",
                      fg_color="transparent", border_width=1,
                      text_color=("gray10", "gray90"),
                      command=self._import_valve_profile_bentley
                      ).grid(row=0, column=1, padx=(0, 6))
        ctk.CTkButton(toolbar, text="Importer DXF",
                      fg_color="transparent", border_width=1,
                      text_color=("gray10", "gray90"),
                      command=self._import_dxf_profile
                      ).grid(row=0, column=2, padx=(0, 6))
        ctk.CTkButton(toolbar, text="Exemple profil",
                      fg_color="transparent", border_width=1,
                      text_color=("gray10", "gray90"),
                      command=self._load_example_profile
                      ).grid(row=0, column=3, padx=(0, 6))

        ctk.CTkLabel(toolbar, text="DN conduite (mm) :",
                     font=ctk.CTkFont(size=12)).grid(row=0, column=4, padx=(12, 4), sticky="e")
        ctk.CTkEntry(toolbar, textvariable=self.var_pipe_dn, width=60
                     ).grid(row=0, column=5, padx=(0, 6))

        ctk.CTkButton(toolbar, text="Calculer ventouses + vidanges",
                      fg_color="#2d6a4f", hover_color="#1b4332",
                      font=ctk.CTkFont(weight="bold"),
                      command=self._run_valve_sizing
                      ).grid(row=0, column=6, padx=(6, 0))

        self.lbl_valve_status = ctk.CTkLabel(toolbar, text="Aucun profil chargé",
                                              text_color="gray", font=ctk.CTkFont(size=11))
        self.lbl_valve_status.grid(row=0, column=7, padx=(12, 0))

        # ── Zone principale : graphique + tableaux ──────────────────
        pane = ctk.CTkFrame(tab)
        pane.grid(row=2, column=0, padx=20, pady=(8, 16), sticky="nsew")
        pane.grid_columnconfigure(0, weight=3)
        pane.grid_columnconfigure(1, weight=2)
        pane.grid_rowconfigure(0, weight=1)

        # Graphique profil en long
        chart_frame = ctk.CTkFrame(pane)
        chart_frame.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="nsew")
        chart_frame.grid_columnconfigure(0, weight=1)
        chart_frame.grid_rowconfigure(0, weight=2)  # Profil (plus grand)
        chart_frame.grid_rowconfigure(1, weight=1)  # Plan (plus petit)

        self.valve_fig = Figure(figsize=(5, 4.5), dpi=100, facecolor="#1a1a2e")
        self.valve_ax = self.valve_fig.add_subplot(211)
        self.valve_canvas = FigureCanvasTkAgg(self.valve_fig, chart_frame)
        self.valve_canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        self.valve_ax.set_xlabel("PK (m)", color="white", fontsize=9)
        self.valve_ax.set_ylabel("Z (m)", color="white", fontsize=9)
        self.valve_ax.set_title("Profil en Long", color="white", fontsize=11)
        self.valve_fig.tight_layout()

        # Graphique tracé en plan (caché par défaut, affiché si DXF chargé)
        self.valve_plan_fig = Figure(figsize=(5, 2.0), dpi=100, facecolor="#1a1a2e")
        self.valve_plan_ax = self.valve_plan_fig.add_subplot(111)
        self.valve_plan_canvas = FigureCanvasTkAgg(self.valve_plan_fig, chart_frame)
        self.valve_plan_canvas.get_tk_widget().grid(row=1, column=0, sticky="nsew")
        self.valve_plan_ax.set_xlabel("X (m)", color="white", fontsize=9)
        self.valve_plan_ax.set_ylabel("Y (m)", color="white", fontsize=9)
        self.valve_plan_ax.set_title("Tracé en Plan (DXF)", color="white", fontsize=10)
        self.valve_plan_ax.text(0.5, 0.5, "Aucun DXF chargé",
                                ha="center", va="center", color="gray",
                                transform=self.valve_plan_ax.transAxes, fontsize=10)
        self.valve_plan_fig.tight_layout()
        self.valve_plan_canvas.draw()

        # Stocker les points du plan DXF
        self._dxf_plan_points: list[tuple[float, float]] = []
        self._dxf_plan_layer: str | None = None

        # Panneau droit : tableaux ventouses + vidanges
        right = ctk.CTkFrame(pane)
        right.grid(row=0, column=1, padx=(5, 10), pady=10, sticky="nsew")
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(1, weight=1)
        right.grid_rowconfigure(3, weight=1)

        ctk.CTkLabel(right, text="Ventouses", font=ctk.CTkFont(size=13, weight="bold")
                     ).grid(row=0, column=0, padx=8, pady=(8, 2), sticky="w")
        self.txt_ventouses = ctk.CTkTextbox(right, font=ctk.CTkFont(family="Consolas", size=11),
                                             state="disabled", height=120)
        self.txt_ventouses.grid(row=1, column=0, padx=8, pady=(0, 8), sticky="nsew")

        ctk.CTkLabel(right, text="Vidanges", font=ctk.CTkFont(size=13, weight="bold")
                     ).grid(row=2, column=0, padx=8, pady=(8, 2), sticky="w")
        self.txt_vidanges = ctk.CTkTextbox(right, font=ctk.CTkFont(family="Consolas", size=11),
                                            state="disabled", height=120)
        self.txt_vidanges.grid(row=3, column=0, padx=8, pady=(0, 8), sticky="nsew")

        ctk.CTkButton(right, text="Exporter CSV",
                      fg_color="transparent", border_width=1,
                      text_color=("gray10", "gray90"),
                      command=self._export_valve_csv
                      ).grid(row=4, column=0, padx=8, pady=(4, 4))

        self.btn_export_vent_report = ctk.CTkButton(
            right, text="📄 Exporter Rapport Ventouses (.docx)",
            fg_color="#1f538d", hover_color="#14406b",
            command=self._export_ventouses_report,
        )
        self.btn_export_vent_report.grid(row=5, column=0, padx=8, pady=(4, 10), sticky="ew")

    # ------------------------------------------------------------------
    # Handlers Phase 3
    # ------------------------------------------------------------------
    def _import_valve_profile(self):
        """Importe un profil en long depuis un CSV."""
        filepath = filedialog.askopenfilename(
            title="Importer profil en long (.csv)",
            filetypes=[("CSV", "*.csv"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        self.valve_profile_filepath = filepath
        ok = self.air_valve_sizer.load_profile_csv(filepath)
        if ok:
            n = len(self.air_valve_sizer.profile)
            self.lbl_valve_status.configure(
                text=f"Profil chargé — {n} points",
                text_color="#2d6a4f")
            self._update_valve_chart()
            self._mark_dirty()
        else:
            messagebox.showerror("Erreur", "Impossible de charger le profil.\n"
                               "Format attendu : CSV avec colonnes PK (m), Z (m).")

    def _import_valve_profile_bentley(self):
        """Importe un profil en long depuis un CSV FlexTable Bentley HAMMER.
        Format : Label, X (m), Y (m), Elevation (m) — distance cumulée calculée
        par cumul de distances successives entre points."""
        filepath = filedialog.askopenfilename(
            title="Importer profil en long (CSV FlexTable Bentley)",
            filetypes=[("CSV FlexTable", "*Profil*.csv *Profile*.csv *Junction*.csv"),
                       ("CSV", "*.csv"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        self.valve_profile_filepath = filepath
        ok = self.air_valve_sizer.load_profile_bentley_csv(filepath)
        if ok:
            n = len(self.air_valve_sizer.profile)
            self.lbl_valve_status.configure(
                text=f"Profil Bentley chargé — {n} points",
                text_color="#2d6a4f")
            self._update_valve_chart()
            self._mark_dirty()
        else:
            messagebox.showerror("Erreur",
                               "Impossible de charger le profil Bentley.\n"
                               "Format attendu : CSV FlexTable avec colonnes\n"
                               "Label, X (m), Y (m), Elevation (m).")

    def _on_column_mapping_request(
        self,
        unknown_col: str,
        available_cols: list[str],
        file_type: str,
        file_hash: str = "",
    ):
        """
        Callback UI appelé par le ColumnMapper quand une colonne n'est pas reconnue.
        Affiche une boîte de dialogue modale et retourne la colonne choisie.
        """
        try:
            return _ask_column_mapping(
                self,
                unknown_col=unknown_col,
                available_cols=available_cols,
                file_type=file_type,
                file_hash=file_hash,
            )
        except Exception as exc:
            print(f"[ColumnMapper] Erreur UI : {exc}")
            return None

    def _import_dxf_profile(self):
        """Importe un profil en long et tracé en plan depuis un DXF.
        Détecte automatiquement les calques 'Tracé en plan' et 'Profil en long'."""
        if not HAS_EZDXF:
            messagebox.showerror("Module manquant",
                               "La librairie ezdxf n'est pas installée.\n"
                               "Lancez : pip install ezdxf")
            return
        filepath = filedialog.askopenfilename(
            title="Importer profil + tracé en plan depuis DXF",
            filetypes=[("DXF (AutoCAD)", "*.dxf"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        result = load_dxf_both(filepath)
        plan_pts = result.get("plan", [])
        prof_pts = result.get("profile", [])

        if not prof_pts:
            messagebox.showerror(
                "Aucun profil trouvé",
                "Calque 'Profil en long' introuvable dans le DXF.\n\n"
                "Calques disponibles :\n" +
                ", ".join(result.get("plan_layer", "") and [result.get("plan_layer")] or [])
            )
            return

        # Charger le profil via load_profile_manual
        self.air_valve_sizer.load_profile_manual([(pk, z) for pk, z in prof_pts])
        self.valve_profile_filepath = filepath

        # Stocker le plan pour affichage secondaire
        self._dxf_plan_points = plan_pts
        self._dxf_plan_layer = result.get("plan_layer")

        n = len(self.air_valve_sizer.profile)
        plan_info = f" + plan ({len(plan_pts)} pts)" if plan_pts else " (pas de plan)"
        self.lbl_valve_status.configure(
            text=f"DXF chargé — {n} points profil{plan_info}",
            text_color="#2d6a4f")
        self._update_valve_chart()
        self._update_valve_plan_chart()
        self._mark_dirty()

    def _load_example_profile(self):
        """Charge un profil exemple pour démonstration."""
        example = [
            (0,    125.5), (100,  130.0), (200,  137.0), (300,  142.0),
            (400,  140.5), (500,  138.0), (600,  135.5), (700,  132.0),
            (800,  130.0), (900,  128.5), (1000, 127.0),
        ]
        self.air_valve_sizer.load_profile_manual(example)
        self.valve_profile_filepath = ""
        n = len(self.air_valve_sizer.profile)
        self.lbl_valve_status.configure(
            text=f"Profil exemple — {n} points",
            text_color="#2d6a4f")
        self._update_valve_chart()
        self._mark_dirty()

    def _run_valve_sizing(self):
        """Lance le dimensionnement ventouses + vidanges."""
        if len(self.air_valve_sizer.profile) < 2:
            messagebox.showwarning("Profil manquant",
                                   "Chargez d'abord un profil en long.")
            return

        try:
            dn = float(self.var_pipe_dn.get())
        except ValueError:
            messagebox.showerror("DN invalide", "Le DN conduite doit être un nombre.")
            return

        self.air_valve_sizer.pipe_dn_mm = dn
        self.air_valve_sizer.size_ventouses()
        self.air_valve_sizer.size_drains()
        self._update_valve_chart()
        self._refresh_valve_textboxes()
        self._mark_dirty()

        nv = len(self.air_valve_sizer.ventouses)
        nd = len(self.air_valve_sizer.vidanges)
        messagebox.showinfo("Dimensionnement terminé",
                            f"Ventouses : {nv}\nVidanges : {nd}")

    def _update_valve_chart(self):
        """Dessine le profil en long avec les ventouses et vidanges."""
        ax = self.valve_ax
        ax.clear()

        profile = self.air_valve_sizer.profile
        if not profile:
            ax.set_xlabel("PK (m)", color="white", fontsize=9)
            ax.set_ylabel("Z (m)", color="white", fontsize=9)
            ax.set_title("Profil en Long", color="white", fontsize=11)
            self.valve_fig.tight_layout()
            self.valve_canvas.draw()
            return

        pks = [p["pk_m"] for p in profile]
        zs = [p["z_m"] for p in profile]

        ax.plot(pks, zs, color="#4cc9f0", linewidth=2, marker="o", markersize=4,
                label="Profil")
        ax.fill_between(pks, zs, min(zs) - 5, alpha=0.15, color="#4cc9f0")

        # Ventouses (▲)
        for v in self.air_valve_sizer.ventouses:
            ax.plot(v["pk_m"], v["z_m"], marker="^", color="#06d6a0", markersize=12,
                    zorder=5)
            ax.annotate(f"V\n{v['type'][:15]}", (v["pk_m"], v["z_m"]),
                        textcoords="offset points", xytext=(0, 10),
                        fontsize=7, color="#06d6a0", ha="center")

        # Vidanges (▼)
        for d in self.air_valve_sizer.vidanges:
            ax.plot(d["pk_m"], d["z_m"], marker="v", color="#ef476f", markersize=12,
                    zorder=5)
            ax.annotate(f"D\n{d['type'][:15]}", (d["pk_m"], d["z_m"]),
                        textcoords="offset points", xytext=(0, -15),
                        fontsize=7, color="#ef476f", ha="center")

        ax.set_xlabel("PK (m)", color="white", fontsize=9)
        ax.set_ylabel("Z (m)", color="white", fontsize=9)
        ax.set_title("Profil en Long", color="white", fontsize=11)
        ax.tick_params(colors="white", labelsize=8)
        for spine in ax.spines.values():
            spine.set_color("#444")
        ax.grid(True, alpha=0.2, color="white")
        ax.legend(loc="upper right", fontsize=8, facecolor="#1a1a2e",
                  edgecolor="#444", labelcolor="white")

        self.valve_fig.tight_layout()
        self.valve_canvas.draw()

    def _update_valve_plan_chart(self):
        """Dessine le tracé en plan à partir des points DXF (si chargés)."""
        ax = self.valve_plan_ax
        ax.clear()

        if not self._dxf_plan_points:
            ax.text(0.5, 0.5, "Aucun DXF chargé",
                    ha="center", va="center", color="gray",
                    transform=ax.transAxes, fontsize=10)
            ax.set_xlabel("X (m)", color="white", fontsize=9)
            ax.set_ylabel("Y (m)", color="white", fontsize=9)
            title = "Tracé en Plan (DXF)"
            ax.set_title(title, color="white", fontsize=10)
            ax.set_facecolor("#1a1a2e")
            for spine in ax.spines.values():
                spine.set_color("#444")
            self.valve_plan_fig.tight_layout()
            self.valve_plan_canvas.draw()
            return

        xs = [p[0] for p in self._dxf_plan_points]
        ys = [p[1] for p in self._dxf_plan_points]
        ax.plot(xs, ys, color="#f72585", linewidth=1.8, marker="o", markersize=3,
                label="Tracé")
        ax.scatter([xs[0]], [ys[0]], color="#06d6a0", s=80, zorder=5,
                   label="Amont", marker="^")
        ax.scatter([xs[-1]], [ys[-1]], color="#ef476f", s=80, zorder=5,
                   label="Aval", marker="v")

        layer = self._dxf_plan_layer or ""
        ax.set_title(f"Tracé en Plan — calque '{layer}'",
                     color="white", fontsize=10)
        ax.set_xlabel("X (m)", color="white", fontsize=9)
        ax.set_ylabel("Y (m)", color="white", fontsize=9)
        ax.tick_params(colors="white", labelsize=8)
        ax.set_aspect("equal", adjustable="datalim")
        for spine in ax.spines.values():
            spine.set_color("#444")
        ax.grid(True, alpha=0.2, color="white")
        ax.legend(loc="upper right", fontsize=7, facecolor="#1a1a2e",
                  edgecolor="#444", labelcolor="white")

        self.valve_plan_fig.tight_layout()
        self.valve_plan_canvas.draw()

    def _refresh_valve_textboxes(self):
        """Met à jour les zones de texte ventouses / vidanges."""
        # Ventouses
        self.txt_ventouses.configure(state="normal")
        self.txt_ventouses.delete("1.0", tk.END)
        if self.air_valve_sizer.ventouses:
            header = f"{'PK (m)':>8}  {'Côte (m)':>9}  {'Type':<35}  {'DN (mm)':>7}\n"
            self.txt_ventouses.insert("1.0", header + "─" * 65 + "\n")
            for v in self.air_valve_sizer.ventouses:
                line = (f"{v['pk_m']:>8.1f}  {v['z_m']:>9.2f}  "
                        f"{v['type']:<35}  {v['dn_mm']:>7}\n")
                self.txt_ventouses.insert(tk.END, line)
        else:
            self.txt_ventouses.insert("1.0", "Aucune ventaise recommandée.\n")
        self.txt_ventouses.configure(state="disabled")

        # Vidanges
        self.txt_vidanges.configure(state="normal")
        self.txt_vidanges.delete("1.0", tk.END)
        if self.air_valve_sizer.vidanges:
            header = (f"{'PK (m)':>8}  {'Côte (m)':>9}  {'Type':<30}  "
                      f"{'DN':>3}  {'Dist.G':>7}  {'Dist.D':>7}\n")
            self.txt_vidanges.insert("1.0", header + "─" * 72 + "\n")
            for d in self.air_valve_sizer.vidanges:
                line = (f"{d['pk_m']:>8.1f}  {d['z_m']:>9.2f}  "
                        f"{d['type']:<30}  {d['dn_mm']:>3}  "
                        f"{d['distance_to_left_m']:>7.1f}  "
                        f"{d['distance_to_right_m']:>7.1f}\n")
                self.txt_vidanges.insert(tk.END, line)
        else:
            self.txt_vidanges.insert("1.0", "Aucune vidange recommandée.\n")
        self.txt_vidanges.configure(state="disabled")

    def _export_valve_csv(self):
        """Exporte les recommandations ventouses + vidanges en CSV."""
        if not self.air_valve_sizer.ventouses and not self.air_valve_sizer.vidanges:
            messagebox.showwarning("Aucune donnée",
                                   "Lancez d'abord le dimensionnement.")
            return
        filepath = filedialog.asksaveasfilename(
            title="Exporter ventouses + vidanges",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")]
        )
        if filepath:
            self.air_valve_sizer.export_csv(filepath)
            messagebox.showinfo("Export réussi", f"Fichier enregistré :\n{filepath}")

    def _export_ventouses_report(self):
        """Exporte un rapport Word (.docx) complet ventouses + vidanges
        avec le profil en long en image."""
        if not VENTOUSES_DOCX_AVAILABLE:
            messagebox.showerror(
                "Module manquant",
                "python-docx n'est pas installé.\n"
                "Lancez : pip install python-docx",
            )
            return
        if not self.air_valve_sizer.profile:
            messagebox.showwarning(
                "Aucun profil",
                "Chargez d'abord un profil en long et lancez le dimensionnement.",
            )
            return
        if not self.air_valve_sizer.ventouses and not self.air_valve_sizer.vidanges:
            messagebox.showwarning(
                "Aucun dimensionnement",
                "Lancez d'abord le calcul des ventouses + vidanges.",
            )
            return

        # Déterminer le nom de fichier par défaut
        default_name = "rapport_ventouses_vidanges.docx"
        if hasattr(self, "entry_projet") and self.entry_projet.get().strip():
            default_name = (
                self.entry_projet.get().strip()
                .replace(" ", "_")
                + "_ventouses.docx"
            )

        filepath = filedialog.asksaveasfilename(
            title="Enregistrer le rapport Ventouses & Vidanges",
            defaultextension=".docx",
            initialfile=default_name,
            filetypes=[("Word Document", "*.docx")],
        )
        if not filepath:
            return

        # Sauvegarder le graphique profil en long en PNG temporaire
        png_tmp = None
        try:
            with tempfile.NamedTemporaryFile(
                suffix=".png", delete=False
            ) as tmp:
                png_tmp = tmp.name
            self.valve_fig.savefig(png_tmp, dpi=120, bbox_inches="tight",
                                    facecolor=self.valve_fig.get_facecolor())
            self.valve_canvas.draw()

            # Construire les métadonnées
            metadata = {
                "nom_projet": (self.entry_projet.get().strip()
                               if hasattr(self, "entry_projet")
                                  and self.entry_projet.get().strip()
                               else "Projet sans nom"),
                "ingenieur":  (self.entry_ingenieur.get().strip()
                               if hasattr(self, "entry_ingenieur")
                                  and self.entry_ingenieur.get().strip()
                               else "—"),
                "date":       (self.entry_date.get().strip()
                               if hasattr(self, "entry_date")
                                  and self.entry_date.get().strip()
                               else ""),
                "dn_mm":      float(self.var_pipe_dn.get())
                               if self.var_pipe_dn.get() else 250.0,
                "profil_source": (self.valve_profile_filepath
                                   or "Profil exemple"),
            }

            # Générer le rapport
            export_ventouses_report(
                air_valve_sizer=self.air_valve_sizer,
                output_path=filepath,
                metadata=metadata,
                profile_chart_png_path=png_tmp,
            )

            messagebox.showinfo(
                "Rapport généré",
                f"Rapport Ventouses & Vidanges enregistré :\n{filepath}",
            )
        except Exception as exc:
            messagebox.showerror(
                "Erreur de génération",
                f"Impossible de générer le rapport :\n{exc}",
            )
        finally:
            # Nettoyer le PNG temporaire
            if png_tmp and os.path.isfile(png_tmp):
                try:
                    os.unlink(png_tmp)
                except OSError:
                    pass

    # ================================================================
    # ACTIONS – Handlers événements
    # ================================================================

    def _get_pn_value(self) -> float:
        return PN_CLASSES.get(self.var_pn.get(), 16.0)

    def _get_pmin_value(self) -> float:
        return PMIN_OPTIONS.get(self.var_pmin.get(), -0.1)

    def _on_pn_change(self, _=None):
        """Mise à jour du graphique quand la PN change."""
        self._update_chart()
        self._update_report_preview()
        self._mark_dirty()

    def _on_pmin_change(self, _=None):
        """Mise à jour du graphique quand la pression min change."""
        self._update_chart()
        self._update_report_preview()
        self._mark_dirty()

    # ------------------------------------------------------------------
    # Gestion des unités d'affichage
    # ------------------------------------------------------------------
    def _on_flow_unit_change(self, _=None):
        """Mise à jour visuelle après changement d'unité de débit."""
        self._update_chart()
        self._update_report_preview()
        # Pas de _mark_dirty : préférence d'affichage, pas une donnée métier.

    def _on_volume_unit_change(self, _=None):
        """Mise à jour visuelle après changement d'unité de volume.
        Met aussi à jour le label d'unité du seuil HPT."""
        self._refresh_threshold_unit_label()
        self._update_chart()
        self._update_report_preview()

    def _on_threshold_change(self, _=None):
        """Mise à jour après édition du seuil HPT (donnée métier → dirty)."""
        self._update_chart()
        self._update_report_preview()
        self._mark_dirty()

    # ------------------------------------------------------------------
    # Helpers de conversion et formatage
    # ------------------------------------------------------------------
    def _format_flow(self, value_m3h):
        """Formate un débit stocké en m³/h (canonique) vers l'unité d'affichage."""
        if value_m3h is None:
            return "—"
        unit = self.var_flow_unit.get()
        factor = FLOW_UNITS.get(unit, 1.0)
        return f"{value_m3h * factor:.2f} {unit}"

    def _format_volume(self, value_l):
        """Formate un volume stocké en L (canonique) vers l'unité d'affichage."""
        if value_l is None:
            return "—"
        unit = self.var_volume_unit.get()
        factor = VOLUME_UNITS.get(unit, 1.0)
        return f"{value_l / factor:.3f} {unit}"

    def _threshold_internal_l(self) -> float:
        """Renvoie la valeur du seuil HPT stockée canoniquement en L."""
        try:
            displayed = float(self.var_volume_threshold.get())
            return displayed * VOLUME_UNITS.get(self.var_volume_unit.get(), 1.0)
        except (ValueError, TypeError):
            return VOLUME_THRESHOLD_L_DEFAULT

    def _volume_threshold_display(self) -> float:
        """Renvoie le seuil HPT dans l'unité d'affichage choisie."""
        return self._threshold_internal_l() / VOLUME_UNITS.get(
            self.var_volume_unit.get(), 1.0)

    def _refresh_threshold_unit_label(self):
        """Met à jour le label '(en L)' ou '(en m³)' à côté du champ seuil."""
        if hasattr(self, "lbl_threshold_unit"):
            self.lbl_threshold_unit.configure(
                text=f"(en {self.var_volume_unit.get()})")

    def _detect_source_unit(self, df, kind: str) -> str:
        """Auto-détecte l'unité source d'un DataFrame à partir de ses colonnes.
        kind = 'flow' ou 'volume'. Retourne l'unité devinée (fallback canonique)."""
        hints = FLOW_UNIT_HINTS if kind == "flow" else VOLUME_UNIT_HINTS
        default = "m³/h" if kind == "flow" else "L"
        if df is None or df.columns is None:
            return default
        for col in df.columns:
            col_lower = str(col).lower()
            for pattern, unit in hints.items():
                if pattern in col_lower:
                    return unit
        return default

    def _resolve_station_src_unit(self, df=None) -> str:
        """Renvoie l'unité source effective (override > auto > canonique)."""
        sel = self.var_station_src_unit.get()
        if sel == "Auto-détection":
            if df is None:
                try:
                    df = self.parser.steady_state_data
                except Exception:
                    df = None
            return self._detect_source_unit(df, "flow")
        return sel

    def _resolve_transient_src_unit(self, df=None) -> str:
        """Renvoie l'unité source effective pour le volume HPT."""
        sel = self.var_transient_src_unit.get()
        if sel == "Auto-détection":
            if df is None:
                try:
                    df = self.parser.hpt_data
                except Exception:
                    df = None
            return self._detect_source_unit(df, "volume")
        return sel

    def _change_appearance_mode(self, mode: str):
        ctk.set_appearance_mode(mode)
        self._update_chart()

    def _show_help(self):
        messagebox.showinfo(
            "Aide — HammerPy Insight v3",
            "WORKFLOW :\n\n"
            "1. Renseignez le nom du projet et l'ingénieur en charge (Onglet 1).\n"
            "2. Choisissez la classe PN de la conduite et la pression min admissible (Sidebar).\n"
            "3. Chargez le fichier station CSV/Excel (Onglet 1 — Régime Permanent).\n"
            "4. Chargez le fichier d'enveloppe HPT CSV/Excel (Onglet 2 — Analyse Transitoire).\n"
            "   → Le graphique interactif avec les seuils critiques s'affiche automatiquement.\n"
            "5. Consultez le rapport dans l'Onglet 3 et exportez en Word (.docx) ou en texte (.txt).\n\n"
            "FORMATS SUPPORTÉS :\n"
            "  • CSV : séparateurs , / ; / Tab – décimales . ou ,\n"
            "  • Excel : .xlsx, .xls\n"
            "  • Encodages : UTF-8, Latin-1, CP1252"
        )

    # ── Import Station ────────────────────────────────────────────────
    def _import_steady_state_file(self):
        filepath = filedialog.askopenfilename(
            title="Sélectionner le fichier de la Station",
            filetypes=[("CSV/Excel", "*.csv *.xlsx *.xls"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        self.station_filepath = filepath
        self.lbl_steady_file.configure(text=os.path.basename(filepath),
                                       text_color=("gray20", "gray80"))
        src_unit = self._resolve_station_src_unit()
        data = self.parser.parse_station_file(filepath, source_unit=src_unit)
        self.steady_state_status = data

        if data["success"]:
            q   = data.get("flow_rate_m3h")
            hmt = data.get("hmt_m")
            self.lbl_flow_val.configure(
                text=self._format_flow(q),
                text_color="#1f8ecf" if q is not None else "orange"
            )
            self.lbl_hmt_val.configure(
                text=f"{hmt} mCE" if hmt is not None else "—",
                text_color="#1f8ecf" if hmt is not None else "orange"
            )
            # Détails des colonnes
            cols_txt = "Colonnes détectées : " + ", ".join(data.get("columns", []))
            self.lbl_steady_details.configure(text=cols_txt[:120] + ("…" if len(cols_txt) > 120 else ""))
            self._update_report_preview()
            self._mark_dirty()
            messagebox.showinfo("Import réussi",
                                f"Fichier chargé : {len(data.get('columns', []))} colonne(s) — {data['n_rows']} ligne(s).\n"
                                f"{'✔ Q et HMT extraits.' if q and hmt else '⚠ Colonnes Q/HMT non trouvées automatiquement.'}")
        else:
            self.lbl_flow_val.configure(text="Erreur", text_color="tomato")
            self.lbl_hmt_val.configure(text="Erreur", text_color="tomato")
            self.lbl_steady_details.configure(text="")
            self._update_report_preview()
            messagebox.showerror("Erreur de chargement",
                                 f"Parsing échoué :\n{data['message']}")

    # ── Import HPT ────────────────────────────────────────────────────
    def _import_transient_file(self):
        filepath = filedialog.askopenfilename(
            title="Sélectionner le fichier transitoire HPT",
            filetypes=[("CSV/Excel", "*.csv *.xlsx *.xls"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        self.hpt_filepath = filepath
        self.lbl_transient_file.configure(text=os.path.basename(filepath),
                                          text_color=("gray20", "gray80"))
        src_unit = self._resolve_transient_src_unit()
        data = self.parser.parse_hpt_file(filepath, source_unit=src_unit)
        self.transient_status = data

        if data["success"]:
            self.lbl_pmin_val.configure(text=f"{data['min_pressure_bar']} bar",
                                        text_color=self._pression_color(data['min_pressure_bar'], "min"))
            self.lbl_pmax_val.configure(text=f"{data['max_pressure_bar']} bar",
                                        text_color=self._pression_color(data['max_pressure_bar'], "max"))
            self.lbl_vgas_val.configure(
                text=self._format_volume(data['max_gas_volume_l']),
                text_color="#33a02c" if data['max_gas_volume_l'] <= 200 else "orange")
            self._update_chart()
            self._update_report_preview()
            self._mark_dirty()
            sim_note = "\n⚠ Les valeurs sont des placeholders (colonnes non détectées)." if data['is_simulated'] else ""
            messagebox.showinfo("Import réussi",
                                f"Fichier chargé : {data['n_rows']} point(s).\n"
                                f"{data['message']}{sim_note}")
        else:
            for lbl in (self.lbl_pmin_val, self.lbl_pmax_val, self.lbl_vgas_val):
                lbl.configure(text="Erreur", text_color="tomato")
            self.lbl_graph_placeholder.configure(
                text="❌  Échec du chargement — Consultez le rapport pour le diagnostic.",
                text_color="tomato"
            )
            if self.canvas:
                self.canvas.get_tk_widget().destroy()
                self.canvas = None
            self.lbl_graph_placeholder.grid(row=0, column=0, padx=20, pady=40, sticky="nsew")
            self._update_report_preview()
            messagebox.showerror("Erreur de chargement", data["message"])

    def _pression_color(self, value: float, kind: str) -> str:
        """Retourne la couleur d'un KPI selon qu'il dépasse la limite ou non."""
        if kind == "min":
            return "#e87c2a" if value < self._get_pmin_value() else "#33a02c"
        if kind == "max":
            return "#e87c2a" if value > self._get_pn_value() else "#33a02c"
        return "#1f8ecf"

    # ------------------------------------------------------------------
    # Import du classeur HAMMER (Flex Tables)
    # ------------------------------------------------------------------
    def _import_workbook(self):
        """Charge un classeur HAMMER (.xlsx/.xls) et affiche le résumé."""
        filepath = filedialog.askopenfilename(
            title="Sélectionner le classeur HAMMER (Flex Tables)",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        ok = self.workbook_manager.load(filepath)
        valid, errors = self.workbook_manager.validate()

        if not valid:
            self.lbl_workbook_file.configure(
                text="❌ " + "; ".join(errors[:2]),
                text_color="tomato"
            )
            messagebox.showerror(
                "Classeur invalide",
                "Le classeur ne contient pas les feuilles obligatoires :\n\n"
                + "\n".join(errors))
            return

        self.workbook_filepath = filepath
        self.lbl_workbook_file.configure(
            text=os.path.basename(filepath),
            text_color=("gray20", "gray80")
        )
        self.btn_reset_workbook.configure(state="normal")

        # Mettre à jour les compteurs
        summary = self.workbook_manager.get_summary()
        for key, lbl in self._wb_counter_labels.items():
            count = summary.get(f"{key}_count", 0)
            lbl.configure(text=str(count) if count > 0 else "—",
                          text_color="#33a02c" if count > 0 else "gray")

        # Mettre à jour les mini-stats
        pmax = summary.get("pmax_bar")
        pmin = summary.get("pmin_bar")
        vmax = summary.get("vmax_pump_ls")
        mats = summary.get("materials", [])

        self._wb_stats["pmax"].configure(
            text=f"{pmax:.2f} bar" if pmax is not None else "— bar",
            text_color="#e87c2a" if pmax and pmax > self._get_pn_value() else "#33a02c")
        self._wb_stats["pmin"].configure(
            text=f"{pmin:.2f} bar" if pmin is not None else "— bar",
            text_color="#e87c2a" if pmin and pmin < self._get_pmin_value() else "#33a02c")
        self._wb_stats["vmax_pump"].configure(
            text=f"{vmax:.1f} L/s" if vmax is not None else "— L/s")
        self._wb_stats["materials"].configure(
            text=", ".join(mats) if mats else "—")

        self._mark_dirty()

        # Avertissements
        warn_msg = ""
        if self.workbook_manager.warnings:
            warn_msg = "\n\nAvertissements :\n" + "\n".join(self.workbook_manager.warnings)

        messagebox.showinfo(
            "Classeur chargé",
            f"Feuilles trouvées : {len(self.workbook_manager.sheet_map)}\n"
            f"Pipes: {summary['pipes_count']} | Nœuds: {summary['nodes_count']} | "
            f"Pompes: {summary['pumps_count']}\n"
            f"Réservoirs: {summary['reservoirs_count']} | HPT: {summary['hpt_count']} | "
            f"Ventouses: {summary['air_valves_count']}"
            + warn_msg
        )

    def _reset_workbook(self):
        """Réinitialise le classeur chargé."""
        self.workbook_manager = WorkbookManager()
        self.workbook_filepath = ""
        self.lbl_workbook_file.configure(text="Aucun classeur chargé", text_color="gray")
        self.btn_reset_workbook.configure(state="disabled")
        for key, lbl in self._wb_counter_labels.items():
            lbl.configure(text="—", text_color="#1f8ecf")
        for key, lbl in self._wb_stats.items():
            unit = {"pmax": "bar", "pmin": "bar", "vmax_pump": "L/s", "materials": ""}.get(key, "")
            lbl.configure(text=f"— {unit}" if unit else "—", text_color="#1f8ecf")
        self._mark_dirty()

    # ------------------------------------------------------------------
    # MULTI-POMPE — Import, sélection, courbe, liste
    # ------------------------------------------------------------------
    PUMP_COLORS = ["#8b5cf6", "#ef4444", "#33a02c", "#e87c2a", "#1f8ecf",
                   "#d946ef", "#06b6d4", "#84cc16"]

    def _selected_pump(self) -> PumpReportParser | None:
        """Retourne la pompe sélectionnée, ou None."""
        i = self.pump_selected_index
        if 0 <= i < len(self.pump_parsers):
            return self.pump_parsers[i]
        return None

    def _import_pump_report(self):
        """Charge un rapport pompe et l'ajoute à la batterie."""
        filepath = filedialog.askopenfilename(
            title="Sélectionner le rapport pompe détaillé",
            filetypes=[("Rapport pompe", "*.rtf *.txt"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        parser = PumpReportParser()
        ok = parser.load(filepath)
        if not ok:
            messagebox.showerror(
                "Rapport invalide",
                "Impossible d'extraire les données pompe :\n\n"
                + "\n".join(parser.errors))
            return

        self.pump_parsers.append(parser)
        self.pump_selected_index = len(self.pump_parsers) - 1

        self._refresh_pump_list()
        self._refresh_pump_kpi()
        self._update_pump_points_display()
        self.btn_remove_pump.configure(state="normal")
        self.btn_add_pump_point.configure(state="normal")
        self.btn_clear_pump_points.configure(state="normal")
        self._update_pump_curve_chart()
        self._mark_dirty()

        summary = parser.get_summary()
        messagebox.showinfo(
            "Pompe ajoutée",
            f"Pompe #{len(self.pump_parsers)} : {summary['label']} (ID {summary['pump_id']})\n"
            f"Débit nominal : {summary['flow_lps']} L/s\n"
            f"HMT pompe : {summary['pump_head_m']} m\n\n"
            f"Total pompes chargées : {len(self.pump_parsers)}"
        )

    def _import_pump_data(self):
        """Importe des données pompe depuis un CSV ou XLSX (Nq, NPSH...)."""
        if not self.pump_parsers:
            messagebox.showwarning(
                "Aucune pompe",
                "Chargez d'abord un ou plusieurs rapports pompe (.rtf)")
            return
        filepath = filedialog.askopenfilename(
            title="Sélectionner le fichier de données pompe (CSV ou XLSX)",
            filetypes=[("Fichier CSV/Excel", "*.csv *.xlsx *.xls"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        ext = os.path.splitext(filepath)[1].lower()
        try:
            if ext in (".xlsx", ".xls"):
                # --- Lecture XLSX ---
                df_raw = pd.read_excel(filepath, engine="openpyxl")
                # Forcer les colonnes en string pour normalisation
                df_raw.columns = [str(c) for c in df_raw.columns]
                df = df_raw.map(lambda x: str(x).replace(",", ".") if isinstance(x, str) else x)
            else:
                # --- Lecture CSV robuste ---
                encodings = ["utf-8", "cp1252", "latin-1", "iso-8859-1"]
                raw = None
                for enc in encodings:
                    try:
                        with open(filepath, "r", encoding=enc) as f:
                            raw = f.read()
                        break
                    except (UnicodeDecodeError, LookupError):
                        continue
                if raw is None:
                    raise ValueError("Encodage non reconnu (essayé utf-8, cp1252, latin-1)")
                first = raw.split("\n")[0].strip()
                sep = ";" if ";" in first else ","
                lines = raw.split("\n")
                header = lines[0]
                data_rows = []
                for line in lines[1:]:
                    if not line.strip():
                        continue
                    parts = line.strip().split(sep)
                    converted = []
                    for i, part in enumerate(parts):
                        if i > 0 and "," in part:
                            part = part.replace(",", ".")
                        converted.append(part)
                    data_rows.append(sep.join(converted))
                csv_text = "\n".join([header] + data_rows)
                df = pd.read_csv(io.StringIO(csv_text), sep=sep, engine="python")
        except Exception as exc:
            messagebox.showerror(
                "Erreur de lecture",
                f"Impossible de lire le fichier :\n{exc}")
            return

        # Normaliser les noms de colonnes
        col_map = {}
        for c in df.columns:
            cl = str(c).strip().lower().replace(" ", "_").replace("-", "_")
            if any(x in cl for x in ("pump_id", "pumpid", "id_pompe")):
                col_map[c] = "pump_id"
            elif cl in ("id", "pump", "label", "nom"):
                col_map[c] = "pump_id"
            elif "nq" in cl:
                col_map[c] = "nq_si"
            elif cl in ("npsh_required_m", "npsh_req", "npsh_requis"):
                col_map[c] = "npsh_required_m"
            elif cl in ("npsh_available_m", "npsh_disp", "npsh_disponible_m"):
                col_map[c] = "npsh_available_m"
        df = df.rename(columns=col_map)

        if "pump_id" not in df.columns:
            messagebox.showerror(
                "Colonne manquante",
                "Le fichier doit contenir une colonne 'pump_id' (ou ID / Pump / Label).\n"
                f"Colonnes trouvées : {', '.join(df.columns)}")
            return

        extra_keys = [c for c in df.columns if c not in ("pump_id",)]
        updated = 0
        errors = []
        for _, row in df.iterrows():
            pid = str(row["pump_id"]).strip()
            matched = False
            for p in self.pump_parsers:
                p_pid = str(p.parsed.get("pump_id", "")).strip()
                p_label = str(p.parsed.get("label", "")).strip()
                if p_pid == pid or p_label == pid:
                    for k in extra_keys:
                        v = row[k]
                        if pd.isna(v):
                            continue
                        try:
                            p.parsed[k] = float(str(v).replace(",", "."))
                        except (ValueError, TypeError):
                            p.parsed[k] = str(v).strip()
                    updated += 1
                    matched = True
                    break
            if not matched:
                errors.append(pid)

        self._refresh_pump_kpi()
        self._mark_dirty()
        msg = f"{updated} pompe(s) mise(s) à jour."
        if errors:
            msg += f"\n\nNon trouvé(es) : {', '.join(errors)}"
        messagebox.showinfo("Import terminé", msg)

    def _generate_templates(self):
        """Génère les 3 fichiers templates XLSX au choix de l'utilisateur."""
        folder = filedialog.askdirectory(
            title="Choisir le dossier de destination des templates")
        if not folder:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Erreur", "openpyxl n'est pas installé.")
            return

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        instr_font = Font(italic=True, color="555555", size=10)

        templates = []

        # ── Template 1 : Données pompe ───────────────────────────────
        wb1 = openpyxl.Workbook()
        ws1 = wb1.active
        ws1.title = "Données pompe"
        headers1 = ["pump_id", "nq_si", "npsh_required_m", "npsh_available_m"]
        for ci, h in enumerate(headers1, 1):
            c = ws1.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        ws1.cell(row=2, column=1, value="PMP-1")
        ws1.cell(row=2, column=2, value=25)
        ws1.cell(row=2, column=3, value=3.5)
        ws1.cell(row=2, column=4, value=15.35)
        ws1.cell(row=3, column=1, value="PMP-2")
        ws1.cell(row=3, column=2, value=25)
        ws1.cell(row=3, column=3, value=3.5)
        ws1.cell(row=3, column=4, value=15.35)
        for ci in range(1, 5):
            ws1.column_dimensions[chr(64 + ci)].width = 20
        # Feuille Instructions
        ws1i = wb1.create_sheet("Instructions")
        ws1i.cell(row=1, column=1,
                  value="Colonnes attendues par l'application :").font = Font(bold=True, size=11)
        for ri, txt in enumerate([
            "pump_id  : Identifiant de la pompe (doit correspondre à l'étiquette dans le rapport)",
            "nq_si    : Vitesse spécifique Nq en unités SI",
            "npsh_required_m : NPSH requis en mètres",
            "npsh_available_m: NPSH disponible en mètres",
            "",
            "Formats acceptés :",
            "- Nombres avec . ou , comme séparateur décimal",
            "- Les colonnes peuvent être dans n'importe quel ordre",
            "- Noms alternatifs reconnus : PumpID, Nq, NPSH_req, NPSH_disp...",
        ], 3):
            ws1i.cell(row=ri, column=1, value=txt).font = instr_font
        ws1i.column_dimensions["A"].width = 60
        templates.append((folder, "template_donnees_pompe.xlsx", wb1))

        # ── Template 2 : Courbe H(Q) ─────────────────────────────────
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Courbe H(Q)"
        headers2 = ["pump_id", "flow_lps", "head_m"]
        for ci, h in enumerate(headers2, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        example = [
            ("PMP-1", 0, 150),
            ("PMP-1", 75, 148),
            ("PMP-1", 150, 130),
            ("PMP-2", 0, 150),
            ("PMP-2", 75, 148),
            ("PMP-2", 150, 130),
        ]
        for ri, (pid, q, h) in enumerate(example, 2):
            ws2.cell(row=ri, column=1, value=pid)
            ws2.cell(row=ri, column=2, value=q)
            ws2.cell(row=ri, column=3, value=h)
        for ci in range(1, 4):
            ws2.column_dimensions[chr(64 + ci)].width = 20
        # Feuille Instructions
        ws2i = wb2.create_sheet("Instructions")
        ws2i.cell(row=1, column=1,
                  value="Colonnes attendues pour la courbe H(Q) :").font = Font(bold=True, size=11)
        for ri, txt in enumerate([
            "pump_id  : Identifiant de la pompe",
            "flow_lps : Débit en L/s (axe X de la courbe)",
            "head_m   : Hauteur manométrique en mètres (axe Y de la courbe)",
            "",
            "Ajoutez autant de points que nécessaire par pompe.",
            "La courbe sera triée automatiquement par débit croissant.",
        ], 3):
            ws2i.cell(row=ri, column=1, value=txt).font = instr_font
        ws2i.column_dimensions["A"].width = 60
        templates.append((folder, "template_courbe_HQ.xlsx", wb2))

        # ── Template 3 : Complet (2 feuilles) ────────────────────────
        wb3 = openpyxl.Workbook()
        # Feuille 1 : Données pompe
        ws3a = wb3.active
        ws3a.title = "Données pompe"
        for ci, h in enumerate(headers1, 1):
            c = ws3a.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        ws3a.cell(row=2, column=1, value="PMP-1")
        ws3a.cell(row=2, column=2, value=25)
        ws3a.cell(row=2, column=3, value=3.5)
        ws3a.cell(row=2, column=4, value=15.35)
        ws3a.cell(row=3, column=1, value="PMP-2")
        ws3a.cell(row=3, column=2, value=25)
        ws3a.cell(row=3, column=3, value=3.5)
        ws3a.cell(row=3, column=4, value=15.35)
        for ci in range(1, 5):
            ws3a.column_dimensions[chr(64 + ci)].width = 20
        # Feuille 2 : Courbe H(Q)
        ws3b = wb3.create_sheet("Courbe H(Q)")
        for ci, h in enumerate(headers2, 1):
            c = ws3b.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        for ri, (pid, q, h) in enumerate(example, 2):
            ws3b.cell(row=ri, column=1, value=pid)
            ws3b.cell(row=ri, column=2, value=q)
            ws3b.cell(row=ri, column=3, value=h)
        for ci in range(1, 4):
            ws3b.column_dimensions[chr(64 + ci)].width = 20
        # Feuille 3 : Instructions
        ws3i = wb3.create_sheet("Instructions")
        ws3i.cell(row=1, column=1,
                  value="Fichier complet — deux feuilles de données :").font = Font(bold=True, size=11)
        for ri, txt in enumerate([
            "Feuille 'Données pompe' : pump_id, nq_si, npsh_required_m, npsh_available_m",
            "Feuille 'Courbe H(Q)'    : pump_id, flow_lps, head_m",
            "",
            "Remplissez les feuilles concernées, enregistrez, puis importez ce fichier",
            "via le bouton 'Importer données pompe (XLSX/CSV)' dans l'application.",
        ], 3):
            ws3i.cell(row=ri, column=1, value=txt).font = instr_font
        ws3i.column_dimensions["A"].width = 70
        templates.append((folder, "template_complet.xlsx", wb3))

        # Sauvegarde
        ok = []
        for fpath, fname, wb in templates:
            try:
                wb.save(os.path.join(fpath, fname))
                ok.append(fname)
            except Exception as exc:
                messagebox.showerror("Erreur", f"Impossible d'écrire {fname} :\n{exc}")
                return
        messagebox.showinfo(
            "Templates générés",
            f"{len(ok)} fichier(s) créé(s) dans :\n{folder}\n\n"
            + "\n".join(f"✓ {f}" for f in ok)
        )


    def _remove_pump(self):
        """Retire la pompe sélectionnée de la batterie."""
        i = self.pump_selected_index
        if i < 0 or i >= len(self.pump_parsers):
            return
        label = self.pump_parsers[i].parsed.get("label", f"Pompe #{i+1}")
        if not messagebox.askyesno("Confirmer",
                                   f"Retirer la pompe « {label} » de la batterie ?"):
            return
        self.pump_parsers.pop(i)
        if self.pump_parsers:
            self.pump_selected_index = min(i, len(self.pump_parsers) - 1)
        else:
            self.pump_selected_index = -1
        self._refresh_pump_list()
        self._refresh_pump_kpi()
        self._update_pump_points_display()
        self._update_pump_curve_chart()
        self.btn_remove_pump.configure(
            state="normal" if self.pump_parsers else "disabled")
        self._mark_dirty()

    def _on_pump_list_click(self, _event=None):
        """Sélection d'une pompe dans la liste."""
        if not self.pump_parsers:
            return
        try:
            index = self.lst_pump_points.index("current")
        except Exception:
            return
        lines = self.lst_pump_points.get("1.0", "end").strip().split("\n")
        if index < len(lines):
            # Extraire le numéro de ligne depuis le texte
            line = lines[index].strip()
            if line and line[0].isdigit():
                idx = int(line.split()[0]) - 1
                if 0 <= idx < len(self.pump_parsers):
                    self.pump_selected_index = idx
                    self._refresh_pump_kpi()
                    self._update_pump_points_display()
                    self._update_pump_curve_chart()

    def _on_pump_mode_change(self):
        """Changement du mode de fonctionnement (Continu / Parallèle)."""
        self._update_pump_curve_chart()
        self._mark_dirty()

    def _refresh_pump_list(self):
        """Rafraîchit la liste texte des pompes chargées."""
        self.lst_pumps.configure(state="normal")
        self.lst_pumps.delete("1.0", "end")
        if not self.pump_parsers:
            self.lst_pumps.insert("1.0", "  (aucune pompe chargée)")
        else:
            header = f"  {'#':>2}   {'Pompe':<12} {'ID':<8} {'Q (L/s)':>8} {'HMT (m)':>8}  {'Mode'}\n"
            self.lst_pumps.insert("1.0", header + "  " + "─" * 60 + "\n")
            for i, p in enumerate(self.pump_parsers, 1):
                s = p.get_summary()
                sel = " ◄" if (i - 1) == self.pump_selected_index else ""
                mode = self.var_pump_mode.get()
                line = (f"  {i:>2}   {s['label']:<12} {str(s['pump_id']):<8} "
                        f"{s['flow_lps'] or 0:>8.1f} {s['pump_head_m'] or 0:>8.1f}  "
                        f"{mode}{sel}\n")
                self.lst_pumps.insert("end", line)
        self.lst_pumps.configure(state="disabled")

    def _refresh_pump_kpi(self):
        """Met à jour les KPI de la pompe sélectionnée."""
        pump = self._selected_pump()
        if pump is None:
            for key, lbl in self._pump_kpi.items():
                lbl.configure(text="—", text_color="#8b5cf6")
            return
        s = pump.get_summary()
        idx = self.pump_selected_index
        self._pump_kpi["idx"].configure(text=str(idx + 1))
        self._pump_kpi["label"].configure(text=s["label"])
        self._pump_kpi["flow"].configure(
            text=f"{s['flow_lps']:.1f}" if s["flow_lps"] is not None else "—")
        self._pump_kpi["head"].configure(
            text=f"{s['pump_head_m']:.1f}" if s["pump_head_m"] is not None else "—")
        npsh_a = s["npsh_available_m"]
        self._pump_kpi["npsh_a"].configure(
            text=f"{npsh_a:.1f}" if npsh_a is not None else "—",
            text_color="#33a02c" if npsh_a is not None and npsh_a > 3 else "#e87c2a")
        self._pump_kpi["n_pts"].configure(text=str(s["n_curve_points"]))

    def _on_add_pump_point(self):
        """Ajoute un point (Q, H) à la courbe de la pompe sélectionnée."""
        pump = self._selected_pump()
        if pump is None:
            messagebox.showwarning("Aucune pompe", "Sélectionnez d'abord une pompe dans la liste.")
            return
        q_str = self.entry_pump_q.get().strip()
        h_str = self.entry_pump_h.get().strip()
        if not q_str or not h_str:
            messagebox.showwarning("Champs requis", "Veuillez saisir Q (L/s) et H (m).")
            return
        try:
            q = float(q_str.replace(",", "."))
            h = float(h_str.replace(",", "."))
        except ValueError:
            messagebox.showerror("Erreur de saisie",
                                 f"Valeurs invalides :\nQ = « {q_str} »\nH = « {h_str} »\n\n"
                                 "Entrez des nombres décimaux (ex: 75 ou 150,5).")
            return
        if q < 0 or h < 0:
            messagebox.showwarning("Valeur négative", "Q et H doivent être ≥ 0.")
            return

        pump.add_curve_point(q, h)
        self.entry_pump_q.delete(0, "end")
        self.entry_pump_h.delete(0, "end")
        self.entry_pump_q.focus_set()
        self._update_pump_points_display()
        self._update_pump_curve_chart()
        self._refresh_pump_kpi()
        self._refresh_pump_list()
        self._mark_dirty()

    def _on_clear_pump_points(self):
        """Efface tous les points de courbe de la pompe sélectionnée."""
        pump = self._selected_pump()
        if pump is None:
            return
        if pump.curve_points:
            if not messagebox.askyesno("Confirmer",
                                       "Effacer tous les points de courbe saisis ?"):
                return
        pump.clear_curve_points()
        self._update_pump_points_display()
        self._update_pump_curve_chart()
        self._refresh_pump_kpi()
        self._refresh_pump_list()
        self._mark_dirty()

    def _update_pump_points_display(self):
        """Rafraîchit la zone texte avec la liste des points de la pompe sélectionnée."""
        self.txt_pump_points.configure(state="normal")
        self.txt_pump_points.delete("1.0", "end")
        pump = self._selected_pump()
        if pump is None or not pump.curve_points:
            self.txt_pump_points.insert("1.0", "  (aucun point saisi)")
        else:
            pts = pump.curve_points
            header = f"  {'#':>3}   {'Q (L/s)':>10}   {'H (m)':>10}\n"
            self.txt_pump_points.insert("1.0", header + "  " + "─" * 30 + "\n")
            for i, p in enumerate(pts, 1):
                line = f"  {i:>3}   {p['flow_lps']:>10.1f}   {p['head_m']:>10.1f}\n"
                self.txt_pump_points.insert("end", line)
        self.txt_pump_points.configure(state="disabled")

    def _update_pump_curve_chart(self):
        """Trace / rafraîchit le graphique H(Q) avec toutes les pompes de la batterie."""
        if self.pump_curve_canvas:
            self.pump_curve_canvas.get_tk_widget().destroy()
            self.pump_curve_canvas = None
        if self.pump_curve_toolbar:
            self.pump_curve_toolbar.destroy()
            self.pump_curve_toolbar = None

        # Vérifier qu'au moins une pompe a ≥ 2 points
        any_curve = any(len(p.curve_points) >= 2 for p in self.pump_parsers)
        if not any_curve:
            self.lbl_pump_curve_placeholder.grid(
                row=0, column=0, padx=20, pady=30, sticky="nsew")
            return

        self.lbl_pump_curve_placeholder.grid_forget()

        import numpy as np
        from matplotlib.figure import Figure

        # Couleurs thème
        mode = ctk.get_appearance_mode()
        if mode == "Dark":
            bg, ax_bg = "#1a1a2e", "#16213e"
            text_c, grid_c, spine_c = "#e0e0e0", "#333355", "#555577"
        else:
            bg, ax_bg = "#f8f9fa", "#ffffff"
            text_c, grid_c, spine_c = "#333333", "#dddddd", "#cccccc"

        fig = Figure(figsize=(7, 3.5), dpi=100)
        fig.patch.set_facecolor(bg)
        ax = fig.add_subplot(111)
        ax.set_facecolor(ax_bg)

        q_global_max = 0
        for idx, pump in enumerate(self.pump_parsers):
            pts = pump.curve_points
            if len(pts) < 2:
                continue
            color = self.PUMP_COLORS[idx % len(self.PUMP_COLORS)]
            q_pts = np.array([p["flow_lps"] for p in pts])
            h_pts = np.array([p["head_m"] for p in pts])

            deg = min(len(pts) - 1, 3)
            coeffs = np.polyfit(q_pts, h_pts, deg)
            poly = np.poly1d(coeffs)

            q_min, q_max = q_pts.min(), q_pts.max()
            margin = max((q_max - q_min) * 0.15, 5.0)
            q_smooth = np.linspace(max(0, q_min - margin), q_max + margin, 200)
            h_smooth = np.maximum(poly(q_smooth), 0)
            q_global_max = max(q_global_max, q_smooth.max())

            label = pump.parsed.get("label", f"Pompe {idx+1}")
            is_selected = (idx == self.pump_selected_index)
            lw = 3.0 if is_selected else 1.8
            alpha = 1.0 if is_selected else 0.6
            ax.plot(q_smooth, h_smooth, color=color, lw=lw, alpha=alpha,
                    label=f"H(Q) {label}", zorder=3 + idx)
            ax.scatter(q_pts, h_pts, color=color, s=60 if is_selected else 40,
                       zorder=5 + idx, edgecolors="white", linewidths=0.8)

            # Point nominal
            flow_nom = pump.parsed.get("flow_lps")
            head_nom = pump.parsed.get("pump_head_m")
            if flow_nom and head_nom:
                ax.scatter([flow_nom], [head_nom], color=color, s=100, zorder=6 + idx,
                           edgecolors="white", linewidths=1.2, marker="D", alpha=alpha)

        # Mode : si parallèle, dessiner la courbe combinée
        if self.var_pump_mode.get() == "Parallèle":
            active = [p for p in self.pump_parsers if len(p.curve_points) >= 2]
            if len(active) > 1:
                # Construire la courbe combinée (somme des débits à chaque H)
                all_h = set()
                for p in active:
                    pts = p.curve_points
                    q_pts = np.array([pp["flow_lps"] for pp in pts])
                    h_pts = np.array([pp["head_m"] for pp in pts])
                    deg = min(len(pts) - 1, 3)
                    coeffs = np.polyfit(q_pts, h_pts, deg)
                    poly = np.poly1d(coeffs)
                    h_min, h_max = h_pts.min(), h_pts.max()
                    all_h.update(np.linspace(max(0, h_min - 10), h_max + 10, 200).tolist())
                all_h = sorted(all_h)
                q_combined = []
                for h_val in all_h:
                    q_total = 0
                    for p in active:
                        pts = p.curve_points
                        q_pts_arr = np.array([pp["flow_lps"] for pp in pts])
                        h_pts_arr = np.array([pp["head_m"] for pp in pts])
                        deg = min(len(pts) - 1, 3)
                        coeffs = np.polyfit(q_pts_arr, h_pts_arr, deg)
                        poly = np.poly1d(coeffs)
                        # Trouver Q pour cette H (inverser la courbe)
                        roots = np.roots(np.append(coeffs, -h_val))
                        real_pos = [r.real for r in roots if abs(r.imag) < 1 and r.real > 0]
                        if real_pos:
                            q_total += min(real_pos)
                    q_combined.append(q_total)
                ax.plot(q_combined, all_h, color="#ffffff", lw=2.5, ls="--",
                        label=f"Combiné ({len(active)} pompes)", zorder=10,
                        alpha=0.8)

        ax.set_xlabel("Débit Q (L/s)", color=text_c, fontweight="bold", fontsize=10)
        ax.set_ylabel("HMT H (m)", color=text_c, fontweight="bold", fontsize=10)
        n = len(self.pump_parsers)
        mode_txt = self.var_pump_mode.get()
        ax.set_title(f"Courbe H(Q) — Batterie ({n} pompe{'s' if n > 1 else ''}, {mode_txt})",
                     color=text_c, fontsize=11, fontweight="bold")
        ax.tick_params(colors=text_c, labelsize=9)
        ax.grid(True, color=grid_c, ls=":", alpha=0.7)
        ax.legend(fontsize=7, loc="upper right", framealpha=0.85,
                  facecolor=ax_bg, labelcolor=text_c, edgecolor=spine_c)
        for sp in ax.spines.values():
            sp.set_color(spine_c)

        fig.tight_layout(pad=1.5)

        self.pump_curve_canvas = FigureCanvasTkAgg(fig, master=self.pump_curve_frame)
        self.pump_curve_canvas.get_tk_widget().grid(
            row=0, column=0, sticky="nsew", padx=4, pady=4)

        toolbar_bg = bg.replace("#", "")
        self.pump_curve_toolbar = tk.Frame(self.pump_curve_frame, bg=toolbar_bg)
        self.pump_curve_toolbar.grid(row=1, column=0, sticky="ew", padx=4)
        self.pump_curve_toolbar = NavigationToolbar2Tk(
            self.pump_curve_canvas, self.pump_curve_toolbar)
        self.pump_curve_toolbar.update()
        self.pump_curve_canvas.draw()

    def _clear_pump_curve_chart(self):
        """Efface le graphique H(Q)."""
        if self.pump_curve_canvas:
            self.pump_curve_canvas.get_tk_widget().destroy()
            self.pump_curve_canvas = None
        if self.pump_curve_toolbar:
            self.pump_curve_toolbar.destroy()
            self.pump_curve_toolbar = None
        self.lbl_pump_curve_placeholder.grid(
            row=0, column=0, padx=20, pady=30, sticky="nsew")

    def _update_chart(self):
        """Trace / rafraîchit le graphique interactif dans l'onglet Analyse Transitoire."""
        # Nettoyage de l'ancien canvas
        if self.canvas:
            self.canvas.get_tk_widget().destroy()
            self.canvas = None
        if self.toolbar:
            self.toolbar.destroy()
            self.toolbar = None
        self.current_fig = None

        if self.parser.hpt_data is None:
            self.lbl_graph_placeholder.grid(row=0, column=0, padx=20, pady=40, sticky="nsew")
            return

        self.lbl_graph_placeholder.grid_forget()

        df   = self.parser.hpt_data
        cols = list(df.columns)

        dist_col = self.parser._find_col(cols, ['distance', 'abscisse', 'position', 'station', 'chainage', 'x (m)'])
        vol_col  = self.parser._find_col(cols, ['volume of gas', 'volume gaz', 'gas volume', 'vol gaz'])
        pmin_col = self.parser._find_col(cols, ['pressure (minimum)', 'pression (minimum)', 'pmin', 'p min',
                                                 'min pressure', 'pressure min', 'pression min'])
        pmax_col = self.parser._find_col(cols, ['pressure (maximum)', 'pression (maximum)', 'pmax', 'p max',
                                                 'max pressure', 'pressure max', 'pression max'])

        x      = df[dist_col] if dist_col else df.index
        y_pmin = df[pmin_col] if pmin_col else ([0] * len(df))
        y_pmax = df[pmax_col] if pmax_col else ([0] * len(df))
        # Volume de gaz : données stockées en L (canonique) → conversion affichage
        vol_unit_factor = VOLUME_UNITS.get(self.var_volume_unit.get(), 1.0)
        y_vgas = (df[vol_col] / vol_unit_factor) if vol_col else ([0] * len(df))
        threshold_disp = self._volume_threshold_display()
        vol_unit_label = self.var_volume_unit.get()

        pn_val   = self._get_pn_value()
        pmin_val = self._get_pmin_value()

        # Palette en fonction du thème
        mode = ctk.get_appearance_mode()
        if mode == "Dark":
            bg, ax_bg = "#2b2b2b", "#1e1e1e"
            text_c, grid_c, spine_c = "#dcdcdc", "#3d3d3d", "#555555"
        else:
            bg, ax_bg = "#ebebeb", "#ffffff"
            text_c, grid_c, spine_c = "#202020", "#dedede", "#bbbbbb"

        fig = Figure(figsize=(8, 4.5), dpi=95)
        fig.patch.set_facecolor(bg)

        ax1 = fig.add_subplot(211)
        ax2 = fig.add_subplot(212, sharex=ax1)

        # ── Pression ────────────────────────────────────────────────
        ax1.set_facecolor(ax_bg)
        ax1.plot(x, y_pmax, color="#e3211b", lw=2.0, label="Pression Max")
        ax1.plot(x, y_pmin, color="#1f78b4", lw=2.0, label="Pression Min")
        ax1.axhline(0,        color="#ff9f00", lw=1.2, ls="--", label="P. Atm. (0 bar)")
        ax1.axhline(pn_val,   color="#ff4444", lw=1.5, ls="-.", alpha=0.8,
                    label=f"Limite PN ({self.var_pn.get()} = {pn_val} bar)")
        ax1.axhline(pmin_val, color="#4488ff", lw=1.5, ls="-.", alpha=0.8,
                    label=f"P min sécurité ({pmin_val} bar)")
        ax1.fill_between(x, y_pmax, pn_val, where=[v > pn_val for v in y_pmax],
                         color="#ff4444", alpha=0.12, label="Zone dépassement PN")
        ax1.fill_between(x, y_pmin, pmin_val, where=[v < pmin_val for v in y_pmin],
                         color="#4488ff", alpha=0.12, label="Zone dépression critique")
        ax1.set_ylabel("Pression (bar)", color=text_c, fontweight="bold", fontsize=9)
        ax1.set_title("Enveloppes de Pression — Coup de Bélier", color=text_c, fontsize=10, fontweight="bold")
        ax1.tick_params(colors=text_c, labelsize=8)
        ax1.grid(True, color=grid_c, ls=":")
        ax1.legend(fontsize=7, loc="upper right", framealpha=0.85,
                   facecolor=ax_bg, labelcolor=text_c, edgecolor=spine_c)
        for sp in ax1.spines.values(): sp.set_color(spine_c)

        # ── Volume de gaz ────────────────────────────────────────────
        ax2.set_facecolor(ax_bg)
        ax2.fill_between(x, y_vgas, color="#33a02c", alpha=0.25)
        ax2.plot(x, y_vgas, color="#33a02c", lw=2.0, label="Vol. Gaz (HPT)")
        ax2.axhline(threshold_disp, color="#ff9f00", lw=1.5, ls="-.", alpha=0.8,
                    label=f"Seuil {threshold_disp:.3f} {vol_unit_label}")
        ax2.set_xlabel("Distance (m)", color=text_c, fontweight="bold", fontsize=9)
        ax2.set_ylabel(f"Vol. Gaz ({vol_unit_label})",
                       color=text_c, fontweight="bold", fontsize=9)
        ax2.set_title("Volume de Gaz Maximum dans le Réservoir HPT", color=text_c, fontsize=10, fontweight="bold")
        ax2.tick_params(colors=text_c, labelsize=8)
        ax2.grid(True, color=grid_c, ls=":")
        ax2.legend(fontsize=7, loc="upper right", framealpha=0.85,
                   facecolor=ax_bg, labelcolor=text_c, edgecolor=spine_c)
        for sp in ax2.spines.values(): sp.set_color(spine_c)

        fig.tight_layout(pad=1.8)
        self.current_fig = fig

        # Intégration dans Tkinter
        self.canvas = FigureCanvasTkAgg(fig, master=self.plot_placeholder_frame)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        self.plot_placeholder_frame.grid_rowconfigure(0, weight=1)
        self.plot_placeholder_frame.grid_rowconfigure(1, weight=0)
        self.plot_placeholder_frame.grid_columnconfigure(0, weight=1)

        toolbar_bg = bg.replace("#", "")
        self.toolbar_frame = tk.Frame(self.plot_placeholder_frame, bg=bg)
        self.toolbar_frame.grid(row=1, column=0, sticky="ew", padx=4)
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        self.toolbar.update()
        self.canvas.draw()

    # ================================================================
    # RAPPORT TEXTUEL (Prévisualisation)
    # ================================================================

    def _update_report_preview(self):
        """Génère et affiche la note technique textuelle dans l'onglet Rapport."""
        projet    = self.entry_projet.get().strip()    if hasattr(self, 'entry_projet')    else ""
        ingenieur = self.entry_ingenieur.get().strip() if hasattr(self, 'entry_ingenieur') else ""
        now       = datetime.now().strftime("%d/%m/%Y  %H:%M")
        pn_lbl    = self.var_pn.get()   if hasattr(self, 'var_pn')   else "PN 16"
        pmin_lbl  = self.var_pmin.get() if hasattr(self, 'var_pmin') else ""
        pn_val    = self._get_pn_value()
        pmin_val  = self._get_pmin_value()

        q   = "--"; hmt  = "--"
        pmax = "--"; pmin_ = "--"; vgas = None
        station_file  = os.path.basename(self.station_filepath) if self.station_filepath else "Non chargé"
        hpt_file      = os.path.basename(self.hpt_filepath)     if self.hpt_filepath     else "Non chargé"

        if self.steady_state_status and self.steady_state_status.get("success"):
            q   = self.steady_state_status.get("flow_rate_m3h") or "--"
            hmt = self.steady_state_status.get("hmt_m")         or "--"
        if self.transient_status and self.transient_status.get("success"):
            pmax  = self.transient_status.get("max_pressure_bar")
            pmin_ = self.transient_status.get("min_pressure_bar")
            vgas  = self.transient_status.get("max_gas_volume_l")

        # Pré-formattage des valeurs avec les unités d'affichage
        flow_disp       = self._format_flow(q)        if q != "--" else "--"
        vgas_disp       = self._format_volume(vgas)   if vgas is not None else "--"
        threshold_disp  = self._volume_threshold_display()
        vol_unit_lbl    = self.var_volume_unit.get()

        lines = [
            "═" * 70,
            "   NOTE TECHNIQUE D'ANALYSE HYDRAULIQUE",
            "   Coup de Bélier — Protection par Réservoir HPT",
            "═" * 70,
            "",
            f"  Projet    : {projet or '(non renseigné)'}",
            f"  Ingénieur : {ingenieur or '(non renseigné)'}",
            f"  Date      : {now}",
            "",
            "─" * 70,
            "  PARAMÈTRES DE DIMENSIONNEMENT (Sidebar)",
            "─" * 70,
            f"  Classe PN de la conduite          : {pn_lbl} → P max admissible = {pn_val} bar",
            f"  Pression minimale de sécurité     : {pmin_val} bar  ({pmin_lbl})",
            f"  Seuil HPT (volume de gaz max)     : {threshold_disp:.3f} {vol_unit_lbl}",
            "",
            "─" * 70,
            "  1. RÉGIME PERMANENT",
            "─" * 70,
            f"  Fichier station          : {station_file}",
            f"  Débit nominal (Q)        : {flow_disp}",
            f"  HMT                      : {hmt} mCE",
            "",
        ]

        # Section Pompe (multi-pompe, si disponible)
        if self.pump_parsers:
            lines += [
                "─" * 70,
                f"  1b. BATTERIE DE POMPES — Mode : {self.var_pump_mode.get()}",
                "─" * 70,
                f"  Nombre de pompes : {len(self.pump_parsers)}",
                "",
            ]
            for i, pump in enumerate(self.pump_parsers, 1):
                s = pump.get_summary()
                pump_file = os.path.basename(pump.filepath) if pump.filepath else "Non chargé"
                pump_flow = f"{s['flow_lps']:.1f} L/s" if s.get("flow_lps") is not None else "—"
                pump_head = f"{s['pump_head_m']:.1f} m" if s.get("pump_head_m") is not None else "—"
                npsh_a = f"{s['npsh_available_m']:.1f} m" if s.get("npsh_available_m") is not None else "—"
                lines += [
                    f"  Pompe {i}/{len(self.pump_parsers)} : {s['label']} (ID {s['pump_id']})",
                    f"    Fichier         : {pump_file}",
                    f"    Conduite aval   : {s.get('downstream_pipe', '—')}",
                    f"    Débit nominal   : {pump_flow}",
                    f"    HMT pompe       : {pump_head}",
                    f"    NPSH disponible : {npsh_a}",
                    f"    Points courbe   : {s.get('n_curve_points', 0)}",
                    "",
                ]

        lines += [
            "─" * 70,
            "  2. ANALYSE TRANSITOIRE — COURBES ENVELOPPES HPT",
            "─" * 70,
            f"  Fichier HPT analysé      : {hpt_file}",
            f"  Pression maximale        : {pmax} bar",
            f"  Pression minimale        : {pmin_} bar",
            f"  Volume de gaz max (HPT)  : {vgas_disp}",
            "",
            "─" * 70,
            "  3. DIAGNOSTIC DE SÉCURITÉ",
            "─" * 70,
        ]

        # Diagnostic Surpression
        if pmax != "--":
            ok = float(pmax) <= pn_val
            icon = "✔" if ok else "⚠"
            lines.append(f"  {icon} Surpression : {pmax} bar  {'<= ' + str(pn_val) + ' bar [OK]' if ok else '> ' + str(pn_val) + ' bar [DÉPASSEMENT PN – ATTENTION]'}")
            if not ok:
                lines.append("    → Augmenter la classe PN ou allonger le temps de fermeture des vannes.")
        else:
            lines.append("  [–] Surpression : données HPT non chargées")

        # Diagnostic Dépression
        if pmin_ != "--":
            ok = float(pmin_) >= pmin_val
            icon = "✔" if ok else "⚠"
            lines.append(f"  {icon} Dépression : {pmin_} bar  {'>= ' + str(pmin_val) + ' bar [OK]' if ok else '< ' + str(pmin_val) + ' bar [DÉPRESSION CRITIQUE]'}")
            if not ok:
                lines.append("    → Vérifier le pré-gonflage HPT. Installer des ventouses automatiques.")
        else:
            lines.append("  [–] Dépression : données HPT non chargées")

        # Diagnostic Volume Gaz (seuil dynamique en unité d'affichage)
        if vgas is not None:
            threshold_l = self._threshold_internal_l()
            ok = float(vgas) <= threshold_l
            icon = "✔" if ok else "⚠"
            lines.append(
                f"  {icon} Volume gaz HPT : {vgas_disp}  "
                f"{'<= ' + f'{threshold_disp:.3f} ' + vol_unit_lbl + ' [OK]' if ok else '> ' + f'{threshold_disp:.3f} ' + vol_unit_lbl + ' [VOLUME ÉLEVÉ]'}")
            if not ok:
                lines.append("    → Augmenter le volume nominal du réservoir HPT.")
        else:
            lines.append("  [–] Volume gaz : données HPT non chargées")

        # Section erreurs de parsing
        errors = []
        if self.steady_state_status and not self.steady_state_status.get("success"):
            errors.append(f"  ✘ Station  : {self.steady_state_status.get('message')}")
        if self.transient_status and not self.transient_status.get("success"):
            errors.append(f"  ✘ HPT      : {self.transient_status.get('message')}")
        if errors:
            lines += ["", "─" * 70, "  ERREURS DE CHARGEMENT", "─" * 70] + errors

        # ── Section Ventaises & Vidanges (Phase 3) ─────────────────
        vents = self.air_valve_sizer.ventouses
        drains = self.air_valve_sizer.vidanges
        if vents or drains:
            lines += [
                "", "─" * 70,
                "  4. PROFIL EN LONG — VENTaises & VIDANGES",
                "─" * 70,
                f"  DN conduite : {self.var_pipe_dn.get()} mm",
                f"  Points profil : {len(self.air_valve_sizer.profile)}",
                "",
            ]
            if vents:
                lines.append(f"  Ventaises recommandées : {len(vents)}")
                lines.append(f"  {'PK (m)':>8}  {'Côte (m)':>9}  {'Type':<35}  {'DN (mm)':>7}")
                lines.append(f"  {'─'*8}  {'─'*9}  {'─'*35}  {'─'*7}")
                for v in vents:
                    lines.append(
                        f"  {v['pk_m']:>8.1f}  {v['z_m']:>9.2f}  "
                        f"{v['type']:<35}  {v['dn_mm']:>7}")
                lines.append("")
            if drains:
                lines.append(f"  Vidanges recommandées : {len(drains)}")
                lines.append(
                    f"  {'PK (m)':>8}  {'Côte (m)':>9}  {'Type':<30}  "
                    f"{'DN':>3}  {'Dist.G':>7}  {'Dist.D':>7}")
                lines.append(
                    f"  {'─'*8}  {'─'*9}  {'─'*30}  {'─'*3}  {'─'*7}  {'─'*7}")
                for d in drains:
                    lines.append(
                        f"  {d['pk_m']:>8.1f}  {d['z_m']:>9.2f}  "
                        f"{d['type']:<30}  {d['dn_mm']:>3}  "
                        f"{d['distance_to_left_m']:>7.1f}  "
                        f"{d['distance_to_right_m']:>7.1f}")
                lines.append("")

        lines += ["", "═" * 70,
                   "  Document généré par HammerPy Insight v3.0",
                  "═" * 70]

        self.txt_report.delete("1.0", tk.END)
        self.txt_report.insert("1.0", "\n".join(lines))

    # ================================================================
    # GESTION DE PROJET (.hpi)  —  Ouvrir / Enregistrer / Enregistrer sous / Quitter
    # ================================================================

    PROJECT_FORMAT_VERSION = "3.0"

    def _update_title(self):
        """Met à jour la barre de titre + l'indicateur d'état en haut de la fenêtre."""
        if self.current_project_path:
            base = os.path.basename(self.current_project_path)
            self.lbl_project_state.configure(text=f"● {base}", text_color="#33a02c")
            self.title(f"HammerPy Insight v3 — {base}"
                       + ("  •  *" if self.is_dirty else ""))
        else:
            if self.is_dirty:
                self.lbl_project_state.configure(
                    text="● Nouveau projet (modifications non sauvegardées)",
                    text_color="orange")
            else:
                self.lbl_project_state.configure(
                    text="● Nouveau projet (non sauvegardé)",
                    text_color="gray")
            self.title("HammerPy Insight v3 — Analyse Transitoire Hydraulique"
                       + ("  •  *" if self.is_dirty else ""))

    def _mark_dirty(self):
        """Marque l'état comme modifié et rafraîchit l'indicateur visuel."""
        if self._suppress_dirty:
            return
        if not self.is_dirty:
            self.is_dirty = True
            self._update_title()

    def _mark_clean(self):
        """Marque l'état comme sauvegardé."""
        self.is_dirty = False
        self._update_title()

    # ------------------------------------------------------------------
    # Phase 4 — Onglet 5 : Système & Diagnostics
    # ------------------------------------------------------------------
    def _setup_diagnostics_tab(self):
        """Crée l'onglet 'Système & Diagnostics' (16 checks A-E)."""
        try:
            from tkinter import ttk
        except ImportError:
            ttk = None
        tab = self.tabview.tab("Système & Diagnostics")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(3, weight=1)

        ctk.CTkLabel(tab, text="Diagnostic Système — 16 vérifications croisées",
                     font=ctk.CTkFont(size=18, weight="bold"), anchor="w"
                     ).grid(row=0, column=0, padx=20, pady=(18, 4), sticky="w")

        ctk.CTkLabel(tab,
                     text="Cohérence Pompe ↔ Réseau ↔ HPT ↔ Ventouses. Catégories A-E.",
                     text_color="gray", anchor="w"
                     ).grid(row=1, column=0, padx=20, pady=(0, 8), sticky="w")

        # ── Bandeau KPI (4 compteurs) ─────────────────────────────────
        kpi_frame = ctk.CTkFrame(tab)
        kpi_frame.grid(row=2, column=0, padx=20, pady=8, sticky="ew")
        kpi_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self.lbl_kpi_ok   = self._make_kpi_card(kpi_frame, "✔ OK",   "0",  "#1a7a1a", col=0)
        self.lbl_kpi_warn = self._make_kpi_card(kpi_frame, "⚠ WARN", "0",  "#c08000", col=1)
        self.lbl_kpi_fail = self._make_kpi_card(kpi_frame, "✘ FAIL", "0",  "#c00000", col=2)
        self.lbl_kpi_na   = self._make_kpi_card(kpi_frame, "— N/A",  "0",  "#888888", col=3)

        # ── Bouton + Treeview ────────────────────────────────────────
        action_frame = ctk.CTkFrame(tab)
        action_frame.grid(row=3, column=0, padx=20, pady=4, sticky="nsew")
        action_frame.grid_columnconfigure(0, weight=1)
        action_frame.grid_rowconfigure(1, weight=1)

        btn_bar = ctk.CTkFrame(action_frame, fg_color="transparent")
        btn_bar.grid(row=0, column=0, padx=8, pady=(8, 4), sticky="ew")
        btn_bar.grid_columnconfigure(0, weight=1)

        ctk.CTkButton(btn_bar, text="  Lancer le diagnostic système  ",
                      command=self._run_diagnostics,
                      font=ctk.CTkFont(size=14, weight="bold")
                      ).grid(row=0, column=0, sticky="w")

        self.lbl_diag_status = ctk.CTkLabel(btn_bar, text="Aucun diagnostic exécuté",
                                            text_color="gray", anchor="w")
        self.lbl_diag_status.grid(row=0, column=1, padx=16, sticky="e")

        # Treeview (tableau plat triable)
        if ttk is not None:
            cols = ("code", "category", "status", "name", "message")
            self.tree_diag = ttk.Treeview(action_frame, columns=cols, show="headings",
                                          height=14, selectmode="browse")
            self.tree_diag.heading("code",     text="Code",        command=lambda: self._sort_tree("code"))
            self.tree_diag.heading("category", text="Catégorie",   command=lambda: self._sort_tree("category"))
            self.tree_diag.heading("status",   text="Statut",      command=lambda: self._sort_tree("status"))
            self.tree_diag.heading("name",     text="Vérification",command=lambda: self._sort_tree("name"))
            self.tree_diag.heading("message",  text="Détail",      command=lambda: self._sort_tree("message"))

            self.tree_diag.column("code",     width=60,  anchor="center")
            self.tree_diag.column("category", width=240, anchor="w")
            self.tree_diag.column("status",   width=80,  anchor="center")
            self.tree_diag.column("name",     width=260, anchor="w")
            self.tree_diag.column("message",  width=520, anchor="w")

            # Style ttk compatible CustomTkinter
            style = ttk.Style()
            try:
                style.theme_use("clam")
            except Exception:
                pass
            style.configure("Treeview", rowheight=24, font=("Segoe UI", 10))
            style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))

            # Couleurs par statut
            self.tree_diag.tag_configure("OK",   foreground="#1a7a1a")
            self.tree_diag.tag_configure("WARN", foreground="#c08000")
            self.tree_diag.tag_configure("FAIL", foreground="#c00000")
            self.tree_diag.tag_configure("NA",   foreground="#888888")

            ysb = ttk.Scrollbar(action_frame, orient="vertical",
                                command=self.tree_diag.yview)
            self.tree_diag.configure(yscrollcommand=ysb.set)
            self.tree_diag.grid(row=1, column=0, padx=(8, 0), pady=(0, 8), sticky="nsew")
            ysb.grid(row=1, column=1, padx=(0, 8), pady=(0, 8), sticky="ns")

            self._diag_sort_state = {"col": "code", "reverse": False}
        else:
            self.tree_diag = None
            ctk.CTkLabel(action_frame, text="(ttk indisponible — diagnostic exécutable via menu)",
                         text_color="gray").grid(row=1, column=0, pady=20)

    def _make_kpi_card(self, parent, title: str, count: str,
                       color: str, col: int) -> ctk.CTkLabel:
        """Crée une carte KPI compacte dans le bandeau."""
        card = ctk.CTkFrame(parent, corner_radius=8)
        card.grid(row=0, column=col, padx=8, pady=8, sticky="ew")
        card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(card, text=title, text_color=color,
                     font=ctk.CTkFont(size=13, weight="bold")
                     ).grid(row=0, column=0, padx=10, pady=(8, 0), sticky="w")
        lbl = ctk.CTkLabel(card, text=count, text_color=color,
                           font=ctk.CTkFont(size=28, weight="bold"))
        lbl.grid(row=1, column=0, padx=10, pady=(0, 8), sticky="w")
        return lbl

    def _run_diagnostics(self):
        """Exécute les 16 checks et peuple le tableau + KPIs."""
        try:
            diag = SystemDiagnostics(
                pump_parsers=self.pump_parsers,
                air_valve_sizer=self.air_valve_sizer if self.air_valve_sizer.profile else None,
                workbook_manager=self.workbook_manager if self.workbook_manager.sheets else None,
                transient_status=self.transient_status,
                pn_value_bar=self._get_pn_value(),
                pmin_value_bar=self._get_pmin_value(),
                vgas_threshold_l=self._threshold_internal_l(),
            )
            checks = diag.run_checks()
            summary = diag.get_summary()
            self.last_diagnostics = checks
            self.last_diagnostics_summary = summary

            # Mise à jour des KPIs
            self.lbl_kpi_ok.configure(text=str(summary.get("OK", 0)))
            self.lbl_kpi_warn.configure(text=str(summary.get("WARN", 0)))
            self.lbl_kpi_fail.configure(text=str(summary.get("FAIL", 0)))
            self.lbl_kpi_na.configure(text=str(summary.get("NA", 0)))

            # Mise à jour du tableau
            if getattr(self, "tree_diag", None) is not None:
                for row in self.tree_diag.get_children():
                    self.tree_diag.delete(row)
                for c in checks:
                    status = c.get("status", "NA")
                    self.tree_diag.insert("", "end", values=(
                        c.get("code", ""),
                        c.get("category", ""),
                        f"{DIAG_ICON.get(status, '?')} {status}",
                        c.get("name", ""),
                        c.get("message", ""),
                    ), tags=(status,))

            # Statut
            n_fail = summary.get("FAIL", 0)
            n_warn = summary.get("WARN", 0)
            n_ok = summary.get("OK", 0)
            self.lbl_diag_status.configure(
                text=f"Diagnostic terminé : {n_ok} OK, {n_warn} WARN, {n_fail} FAIL",
                text_color="#c00000" if n_fail else ("#c08000" if n_warn else "#1a7a1a"),
            )
            self._mark_dirty()
        except Exception as exc:
            self.lbl_diag_status.configure(
                text=f"Erreur : {exc}", text_color="#c00000")
            messagebox.showerror("Erreur diagnostic système", str(exc))

    def _sort_tree(self, col: str):
        """Tri du Treeview par colonne."""
        if getattr(self, "tree_diag", None) is None:
            return
        state = self._diag_sort_state
        reverse = False
        if state["col"] == col:
            reverse = not state["reverse"]
        state["col"] = col
        state["reverse"] = reverse

        rows = [(self.tree_diag.set(iid, col), iid)
                for iid in self.tree_diag.get_children("")]
        # Tri naturel
        def _key(item):
            v = item[0]
            # Essayer numérique
            try:
                return (0, float(v))
            except (ValueError, TypeError):
                return (1, v.lower())
        rows.sort(key=_key, reverse=reverse)
        for idx, (_, iid) in enumerate(rows):
            self.tree_diag.move(iid, "", idx)

    def _refresh_diagnostics_table(self):
        """Recharge le tableau + KPI depuis last_diagnostics (chargement projet)."""
        if not self.last_diagnostics:
            return
        summary = self.last_diagnostics_summary or {}
        if hasattr(self, "lbl_kpi_ok"):
            self.lbl_kpi_ok.configure(text=str(summary.get("OK", 0)))
            self.lbl_kpi_warn.configure(text=str(summary.get("WARN", 0)))
            self.lbl_kpi_fail.configure(text=str(summary.get("FAIL", 0)))
            self.lbl_kpi_na.configure(text=str(summary.get("NA", 0)))
        if getattr(self, "tree_diag", None) is not None:
            for row in self.tree_diag.get_children():
                self.tree_diag.delete(row)
            for c in self.last_diagnostics:
                status = c.get("status", "NA")
                self.tree_diag.insert("", "end", values=(
                    c.get("code", ""),
                    c.get("category", ""),
                    f"{DIAG_ICON.get(status, '?')} {status}",
                    c.get("name", ""),
                    c.get("message", ""),
                ), tags=(status,))
        if hasattr(self, "lbl_diag_status"):
            n_fail = summary.get("FAIL", 0)
            n_warn = summary.get("WARN", 0)
            n_ok = summary.get("OK", 0)
            self.lbl_diag_status.configure(
                text=f"Diagnostic chargé : {n_ok} OK, {n_warn} WARN, {n_fail} FAIL",
                text_color="#c00000" if n_fail else ("#c08000" if n_warn else "#1a7a1a"),
            )

    # ------------------------------------------------------------------
    # Construction de l'état complet du projet pour sérialisation
    # ------------------------------------------------------------------
    def _build_project_payload(self) -> dict:
        """Sérialise l'intégralité de l'état de la session en dictionnaire."""
        try:
            station_data = (
                self.parser.steady_state_data.to_dict(orient="records")
                if self.parser.steady_state_data is not None else None
            )
        except Exception:
            station_data = None

        try:
            hpt_data = (
                self.parser.hpt_data.to_dict(orient="records")
                if self.parser.hpt_data is not None else None
            )
        except Exception:
            hpt_data = None

        # Données du classeur HAMMER (6 feuilles, orient='records' pour compacité)
        workbook_data = {}
        for canonical in ["pipes", "nodes", "reservoirs", "pumps", "hpt", "air_valves"]:
            df = self.workbook_manager.sheets.get(canonical)
            if df is not None and len(df) > 0:
                try:
                    # Convertir les types numpy en types natifs Python pour JSON
                    records = df.to_dict(orient="records")
                    for rec in records:
                        for k, v in rec.items():
                            if hasattr(v, 'item'):
                                rec[k] = v.item()
                    workbook_data[canonical] = records
                except Exception:
                    workbook_data[canonical] = []

        return {
            "version": self.PROJECT_FORMAT_VERSION,
            "metadata": {
                "nom_projet":       self.entry_projet.get().strip()
                                     if hasattr(self, "entry_projet") else "",
                "ingenieur":        self.entry_ingenieur.get().strip()
                                     if hasattr(self, "entry_ingenieur") else "",
                "date_creation":    self.project_creation_date,
                "date_modification": datetime.now().isoformat(timespec="seconds"),
            },
            "config": {
                "pn_label":   self.var_pn.get(),
                "pn_value":   self._get_pn_value(),
                "pmin_label": self.var_pmin.get(),
                "pmin_value": self._get_pmin_value(),
                "flow_unit":             self.var_flow_unit.get(),
                "volume_unit":           self.var_volume_unit.get(),
                "volume_threshold_l":    self._threshold_internal_l(),
            },
            "station": {
                "filepath": self.station_filepath,
                "data":     station_data,
                "status":   self.steady_state_status,
            },
            "hpt": {
                "filepath": self.hpt_filepath,
                "data":     hpt_data,
                "status":   self.transient_status,
            },
            "workbook": {
                "filepath": self.workbook_filepath,
                "sheets":   workbook_data,
            },
            "pump_group": {
                "mode": self.var_pump_mode.get(),
                "pumps": [
                    {
                        "filepath":     p.filepath,
                        "parsed":       p.parsed,
                        "curve_points": p.curve_points,
                    }
                    for p in self.pump_parsers
                ],
                "selectedIndex": self.pump_selected_index,
            },
            "report_text": (self.txt_report.get("1.0", tk.END)
                            if hasattr(self, "txt_report") else ""),

            "air_valves": {
                "filepath":   self.valve_profile_filepath,
                "pipe_dn_mm": float(self.var_pipe_dn.get())
                              if self.var_pipe_dn.get() else 250.0,
                "profile":    self.air_valve_sizer.profile,
                "ventouses":  self.air_valve_sizer.ventouses,
                "vidanges":   self.air_valve_sizer.vidanges,
                "dxf_plan_points": self._dxf_plan_points,
                "dxf_plan_layer":  self._dxf_plan_layer,
            },

            "column_mappings": self._column_mapper.serialize(),

            "system_diagnostics": {
                "checks":  self.last_diagnostics or [],
                "summary": self.last_diagnostics_summary or {},
            } if self.last_diagnostics else {},
        }

    # ------------------------------------------------------------------
    # Sauvegarde / chargement bas niveau
    # ------------------------------------------------------------------
    def _save_project(self, filepath: str):
        """Sérialise l'état du projet dans un fichier .hpi (JSON)."""
        payload = self._build_project_payload()
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        self.current_project_path = filepath
        self._mark_clean()
        messagebox.showinfo(
            "Projet enregistré",
            f"Projet sauvegardé avec succès :\n{filepath}")

    def _load_project(self, filepath: str):
        """Charge un fichier .hpi et restaure l'état complet de l'application."""
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                payload = json.load(f)
        except json.JSONDecodeError as exc:
            messagebox.showerror("Fichier invalide",
                                 f"Le fichier .hpi est corrompu ou non lisible :\n{exc}")
            return
        except Exception as exc:
            messagebox.showerror("Erreur de chargement", str(exc))
            return

        version = payload.get("version", "1.0")
        if version not in ("2.0", "3.0"):
            messagebox.showwarning(
                "Version incompatible",
                f"Ce projet a été créé avec une version différente "
                f"({version}).\nLa tentative de chargement peut échouer.")

        # Anti-boucle : empêche les handlers de marquer dirty pendant la restauration
        self._suppress_dirty = True
        try:
            # ── Métadonnées ──────────────────────────────────────────
            meta = payload.get("metadata", {})
            self.project_creation_date = (meta.get("date_creation")
                                          or datetime.now().isoformat(timespec="seconds"))
            if hasattr(self, "entry_projet"):
                self.entry_projet.delete(0, tk.END)
                self.entry_projet.insert(0, meta.get("nom_projet", ""))
            if hasattr(self, "entry_ingenieur"):
                self.entry_ingenieur.delete(0, tk.END)
                self.entry_ingenieur.insert(0, meta.get("ingenieur", ""))

            # ── Configuration PN / Pmin ──────────────────────────────
            cfg = payload.get("config", {})
            if "pn_label" in cfg and cfg["pn_label"] in PN_CLASSES:
                self.var_pn.set(cfg["pn_label"])
            if "pmin_label" in cfg and cfg["pmin_label"] in PMIN_OPTIONS:
                self.var_pmin.set(cfg["pmin_label"])
            # Unités d'affichage
            if "flow_unit" in cfg and cfg["flow_unit"] in FLOW_UNITS:
                self.var_flow_unit.set(cfg["flow_unit"])
            if "volume_unit" in cfg and cfg["volume_unit"] in VOLUME_UNITS:
                self.var_volume_unit.set(cfg["volume_unit"])
            # Seuil HPT : stocké en L, on l'affiche dans l'unité choisie
            if "volume_threshold_l" in cfg:
                threshold_disp = cfg["volume_threshold_l"] / VOLUME_UNITS.get(
                    self.var_volume_unit.get(), 1.0)
                self.var_volume_threshold.set(f"{threshold_disp:.3f}")
            self._refresh_threshold_unit_label()

            # ── Station (régime permanent) ───────────────────────────
            station = payload.get("station", {})
            self.station_filepath = station.get("filepath", "") or ""
            self.steady_state_status = station.get("status")
            try:
                self.parser.steady_state_data = (
                    pd.DataFrame.from_records(station["data"])
                    if station.get("data") is not None else None
                )
            except Exception:
                self.parser.steady_state_data = None

            # Mise à jour visuelle des KPI station
            self.lbl_steady_file.configure(
                text=os.path.basename(self.station_filepath)
                     if self.station_filepath else "Aucun fichier sélectionné",
                text_color=("gray20", "gray80")
                          if self.station_filepath else "gray")
            s_ok = (self.steady_state_status or {}).get("success")
            if s_ok:
                q   = self.steady_state_status.get("flow_rate_m3h")
                hmt = self.steady_state_status.get("hmt_m")
                self.lbl_flow_val.configure(
                    text=f"{q} m³/h" if q is not None else "—",
                    text_color="#1f8ecf" if q is not None else "orange")
                self.lbl_hmt_val.configure(
                    text=f"{hmt} mCE" if hmt is not None else "—",
                    text_color="#1f8ecf" if hmt is not None else "orange")
                cols = self.steady_state_status.get("columns", [])
                self.lbl_steady_details.configure(
                    text=("Colonnes détectées : " + ", ".join(cols))[:160])
            else:
                self.lbl_flow_val.configure(text="-- --", text_color="#1f8ecf")
                self.lbl_hmt_val.configure(text="-- --", text_color="#1f8ecf")
                self.lbl_steady_details.configure(text="")

            # ── HPT (transitoire) ────────────────────────────────────
            hpt = payload.get("hpt", {})
            self.hpt_filepath = hpt.get("filepath", "") or ""
            self.transient_status = hpt.get("status")
            try:
                self.parser.hpt_data = (
                    pd.DataFrame.from_records(hpt["data"])
                    if hpt.get("data") is not None else None
                )
            except Exception:
                self.parser.hpt_data = None

            self.lbl_transient_file.configure(
                text=os.path.basename(self.hpt_filepath)
                     if self.hpt_filepath else "Aucun fichier HPT chargé",
                text_color=("gray20", "gray80")
                          if self.hpt_filepath else "gray")
            t_ok = (self.transient_status or {}).get("success")
            if t_ok:
                self.lbl_pmin_val.configure(
                    text=f"{self.transient_status.get('min_pressure_bar')} bar",
                    text_color=self._pression_color(
                        self.transient_status.get("min_pressure_bar"), "min"))
                self.lbl_pmax_val.configure(
                    text=f"{self.transient_status.get('max_pressure_bar')} bar",
                    text_color=self._pression_color(
                        self.transient_status.get("max_pressure_bar"), "max"))
                vgas = self.transient_status.get("max_gas_volume_l")
                self.lbl_vgas_val.configure(
                    text=f"{vgas} L",
                    text_color="#33a02c" if vgas is not None and vgas <= 200 else "orange")
            else:
                self.lbl_pmin_val.configure(text="-- bar", text_color="#1f8ecf")
                self.lbl_pmax_val.configure(text="-- bar", text_color="#1f8ecf")
                self.lbl_vgas_val.configure(text="-- L", text_color="#1f8ecf")

            # ── Classeur HAMMER (v3.0+) ───────────────────────────────
            wb = payload.get("workbook", {})
            self.workbook_filepath = wb.get("filepath", "") or ""
            sheets_data = wb.get("sheets", {})

            if sheets_data:
                # Reconstruire les DataFrames à partir des records JSON
                self.workbook_manager = WorkbookManager()
                self.workbook_manager.filepath = self.workbook_filepath
                for canonical, records in sheets_data.items():
                    if records:
                        try:
                            self.workbook_manager.sheets[canonical] = pd.DataFrame.from_records(records)
                        except Exception:
                            self.workbook_manager.sheets[canonical] = pd.DataFrame()

                # Mettre à jour l'UI
                self.lbl_workbook_file.configure(
                    text=os.path.basename(self.workbook_filepath)
                         if self.workbook_filepath else "Aucun classeur chargé",
                    text_color=("gray20", "gray80")
                              if self.workbook_filepath else "gray")
                self.btn_reset_workbook.configure(
                    state="normal" if self.workbook_filepath else "disabled")

                summary = self.workbook_manager.get_summary()
                for key, lbl in self._wb_counter_labels.items():
                    count = summary.get(f"{key}_count", 0)
                    lbl.configure(text=str(count) if count > 0 else "—",
                                  text_color="#33a02c" if count > 0 else "gray")

                pmax = summary.get("pmax_bar")
                pmin = summary.get("pmin_bar")
                vmax = summary.get("vmax_pump_ls")
                mats = summary.get("materials", [])
                self._wb_stats["pmax"].configure(
                    text=f"{pmax:.2f} bar" if pmax is not None else "— bar",
                    text_color="#e87c2a" if pmax and pmax > self._get_pn_value() else "#33a02c")
                self._wb_stats["pmin"].configure(
                    text=f"{pmin:.2f} bar" if pmin is not None else "— bar",
                    text_color="#e87c2a" if pmin and pmin < self._get_pmin_value() else "#33a02c")
                self._wb_stats["vmax_pump"].configure(
                    text=f"{vmax:.1f} L/s" if vmax is not None else "— L/s")
                self._wb_stats["materials"].configure(
                    text=", ".join(mats) if mats else "—")
            else:
                self._reset_workbook()

            # ── Pompe (multi-pompe v3.1+ / rétrocompatible v3.0) ─────
            pump_group = payload.get("pump_group", {})
            pump_legacy = payload.get("pump", {})  # rétrocompatible v3.0
            self.pump_parsers = []
            self.pump_selected_index = -1

            if pump_group.get("pumps"):
                # Format v3.1+ : liste de pompes
                if "mode" in pump_group:
                    self.var_pump_mode.set(pump_group["mode"])
                for p_data in pump_group["pumps"]:
                    if p_data.get("parsed"):
                        p = PumpReportParser()
                        p.filepath = p_data.get("filepath", "")
                        p.parsed = p_data["parsed"]
                        p.curve_points = p_data.get("curve_points", [])
                        self.pump_parsers.append(p)
                self.pump_selected_index = pump_group.get("selectedIndex", 0)
                if self.pump_parsers:
                    self.pump_selected_index = min(
                        self.pump_selected_index, len(self.pump_parsers) - 1)
            elif pump_legacy.get("parsed"):
                # Format v3.0 : une seule pompe
                p = PumpReportParser()
                p.filepath = pump_legacy.get("filepath", "")
                p.parsed = pump_legacy["parsed"]
                p.curve_points = pump_legacy.get("curve_points", [])
                self.pump_parsers = [p]
                self.pump_selected_index = 0

            self._refresh_pump_list()
            self._refresh_pump_kpi()
            self._update_pump_points_display()
            self.btn_remove_pump.configure(
                state="normal" if self.pump_parsers else "disabled")
            self.btn_add_pump_point.configure(
                state="normal" if self.pump_parsers else "disabled")
            self.btn_clear_pump_points.configure(
                state="normal" if self.pump_parsers else "disabled")

            # ── Ventouses & Vidanges (Phase 3, rétrocompatible v3.0) ──
            air_data = payload.get("air_valves", {})
            if air_data:
                self.valve_profile_filepath = air_data.get("filepath", "") or ""
                dn = air_data.get("pipe_dn_mm", 250.0)
                self.var_pipe_dn.set(str(int(dn)) if dn == int(dn) else str(dn))
                self.air_valve_sizer.pipe_dn_mm = float(dn)
                self.air_valve_sizer.profile = air_data.get("profile", [])
                self.air_valve_sizer.ventouses = air_data.get("ventouses", [])
                self.air_valve_sizer.vidanges = air_data.get("vidanges", [])
                self._dxf_plan_points = air_data.get("dxf_plan_points", []) or []
                self._dxf_plan_layer = air_data.get("dxf_plan_layer")
                self._update_valve_chart()
                self._update_valve_plan_chart()
                self._refresh_valve_textboxes()
                n = len(self.air_valve_sizer.profile)
                plan_info = f" + {len(self._dxf_plan_points)} pts plan" if self._dxf_plan_points else ""
                self.lbl_valve_status.configure(
                    text=f"Projet chargé — {n} points profil{plan_info}" if n else "Aucun profil",
                    text_color="#2d6a4f" if n else "gray")

            # ── Column Mappings (Phase 3.5+, rétrocompatible v3.0) ───
            mappings_data = payload.get("column_mappings", {})
            if mappings_data:
                self._column_mapper.deserialize(mappings_data)

            # ── System Diagnostics (Phase 4, rétrocompatible v3.0) ────
            diag_data = payload.get("system_diagnostics", {})
            if diag_data:
                self.last_diagnostics = diag_data.get("checks") or []
                self.last_diagnostics_summary = diag_data.get("summary") or None
                if self.last_diagnostics and getattr(self, "tree_diag", None) is not None:
                    self._refresh_diagnostics_table()

            # ── Rapport ──────────────────────────────────────────────
            report_text = payload.get("report_text", "")
            if hasattr(self, "txt_report") and report_text:
                self.txt_report.delete("1.0", tk.END)
                self.txt_report.insert("1.0", report_text)

            # ── Graphique (régénéré à partir des données) ───────────
            self._update_chart()
            self._update_report_preview()

        finally:
            self._suppress_dirty = False

        self.current_project_path = filepath
        self._mark_clean()
        messagebox.showinfo(
            "Projet chargé",
            f"Projet ouvert avec succès :\n{filepath}")

    # ------------------------------------------------------------------
    # Commandes du menu (boutons)
    # ------------------------------------------------------------------
    def _open_project(self):
        """Ouvre un fichier .hpi existant après confirmation éventuelle."""
        if not self._confirm_discard_changes("ouvrir un autre projet"):
            return
        filepath = filedialog.askopenfilename(
            title="Ouvrir un projet HammerPy Insight",
            defaultextension=".hpi",
            filetypes=[("Projet HammerPy Insight", "*.hpi"),
                       ("Tous les fichiers", "*.*")]
        )
        if not filepath:
            return
        self._load_project(filepath)

    def _save_project_as(self):
        """Enregistre le projet sous un nouveau chemin."""
        default_name = "projet_hammerpy.hpi"
        if hasattr(self, "entry_projet"):
            proj = self.entry_projet.get().strip()
            if proj:
                default_name = (proj.replace(" ", "_").replace("/", "-")
                                + ".hpi")
        filepath = filedialog.asksaveasfilename(
            title="Enregistrer le projet sous…",
            defaultextension=".hpi",
            initialfile=default_name,
            filetypes=[("Projet HammerPy Insight", "*.hpi"),
                       ("Tous les fichiers", "*.*")]
        )
        if not filepath:
            return
        try:
            self._save_project(filepath)
        except Exception as exc:
            messagebox.showerror("Erreur d'enregistrement", str(exc))

    def _save_project_quick(self):
        """Enregistre rapidement (au dernier chemin connu, ou bascule sur 'Enregistrer sous')."""
        if self.current_project_path and os.path.exists(self.current_project_path):
            try:
                payload = self._build_project_payload()
                with open(self.current_project_path, "w", encoding="utf-8") as f:
                    json.dump(payload, f, indent=2, ensure_ascii=False)
                self._mark_clean()
                messagebox.showinfo(
                    "Projet enregistré",
                    f"Projet sauvegardé :\n{self.current_project_path}")
            except Exception as exc:
                messagebox.showerror("Erreur d'enregistrement", str(exc))
        else:
            self._save_project_as()

    def _quit_app(self):
        """Quitte l'application après confirmation éventuelle."""
        if self.is_dirty:
            choice = messagebox.askyesnocancel(
                "Modifications non sauvegardées",
                "Le projet courant contient des modifications non sauvegardées.\n\n"
                "Voulez-vous les enregistrer avant de quitter ?\n\n"
                "• Oui  : enregistrer puis quitter\n"
                "• Non  : quitter sans enregistrer\n"
                "• Annuler : revenir à l'application")
            if choice is None:   # Annuler
                return
            if choice:            # Oui → enregistrer
                if self.current_project_path and os.path.exists(self.current_project_path):
                    try:
                        payload = self._build_project_payload()
                        with open(self.current_project_path, "w", encoding="utf-8") as f:
                            json.dump(payload, f, indent=2, ensure_ascii=False)
                    except Exception as exc:
                        messagebox.showerror(
                            "Erreur d'enregistrement",
                            f"Impossible d'enregistrer avant de quitter :\n{exc}")
                        return
                else:
                    filepath = filedialog.asksaveasfilename(
                        title="Enregistrer le projet avant de quitter",
                        defaultextension=".hpi",
                        initialfile="projet_hammerpy.hpi",
                        filetypes=[("Projet HammerPy Insight", "*.hpi")]
                    )
                    if not filepath:
                        return
                    try:
                        self._save_project(filepath)
                    except Exception as exc:
                        messagebox.showerror("Erreur d'enregistrement", str(exc))
                        return
        self.destroy()

    def _confirm_discard_changes(self, action_label: str) -> bool:
        """Demande confirmation si modifications non sauvegardées. Retourne True si OK."""
        if not self.is_dirty:
            return True
        choice = messagebox.askyesnocancel(
            "Modifications non sauvegardées",
            f"Vous allez {action_label}, mais le projet courant contient "
            f"des modifications non sauvegardées.\n\n"
            f"Voulez-vous les enregistrer d'abord ?\n\n"
            f"• Oui  : enregistrer puis continuer\n"
            f"• Non  : continuer sans enregistrer (perte des modifications)\n"
            f"• Annuler : revenir à l'application")
        if choice is None:
            return False
        if choice:
            if self.current_project_path and os.path.exists(self.current_project_path):
                try:
                    payload = self._build_project_payload()
                    with open(self.current_project_path, "w", encoding="utf-8") as f:
                        json.dump(payload, f, indent=2, ensure_ascii=False)
                    self._mark_clean()
                except Exception as exc:
                    messagebox.showerror(
                        "Erreur d'enregistrement",
                        f"Impossible d'enregistrer : {exc}\nOpération annulée.")
                    return False
            else:
                filepath = filedialog.asksaveasfilename(
                    title="Enregistrer le projet",
                    defaultextension=".hpi",
                    initialfile="projet_hammerpy.hpi",
                    filetypes=[("Projet HammerPy Insight", "*.hpi")]
                )
                if not filepath:
                    return False
                try:
                    self._save_project(filepath)
                except Exception as exc:
                    messagebox.showerror("Erreur d'enregistrement", str(exc))
                    return False
        return True

    # ================================================================
    # EXPORT
    # ================================================================

    def _export_txt(self):
        """Export de la note textuelle brute au format .txt."""
        filepath = filedialog.asksaveasfilename(
            title="Exporter la note (.txt)",
            defaultextension=".txt",
            filetypes=[("Texte", "*.txt"), ("Tous", "*.*")]
        )
        if filepath:
            try:
                with open(filepath, "w", encoding="utf-8") as f:
                    f.write(self.txt_report.get("1.0", tk.END))
                messagebox.showinfo("Export réussi", f"Fichier enregistré :\n{filepath}")
            except Exception as exc:
                messagebox.showerror("Erreur", str(exc))

    def _export_word(self):
        """Export du rapport professionnel au format Word .docx avec graphique intégré."""
        if not DOCX_AVAILABLE:
            messagebox.showerror("python-docx manquant",
                                 "Installez python-docx :\n  pip install python-docx")
            return

        filepath = filedialog.asksaveasfilename(
            title="Enregistrer le rapport Word",
            defaultextension=".docx",
            filetypes=[("Word", "*.docx"), ("Tous", "*.*")]
        )
        if not filepath:
            return

        # ── Sauvegardes temporaires des PNG (graphique HPT + 4 graphes diag) ─
        # Liste de tous les PNG temporaires à nettoyer dans le finally:
        chart_paths: list[str | None] = []

        # 1. Graphique HPT (existant)
        chart_path = None
        if self.current_fig:
            try:
                tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                self.current_fig.savefig(tmp.name, dpi=150, bbox_inches="tight",
                                         facecolor=self.current_fig.get_facecolor())
                tmp.close()
                chart_path = tmp.name
            except Exception:
                chart_path = None
        chart_paths.append(chart_path)

        # 2-5. Graphiques Phase 4 (diagnostics_charts)
        from diagnostics_charts import (
            build_kpi_donut, build_category_stack,
            build_compliance_bars, build_profile_chart,
        )
        diag_kpi_path        = None
        diag_category_path   = None
        diag_compliance_path = None
        diag_profile_path    = None
        try:
            diag_kpi_path = build_kpi_donut(self.last_diagnostics_summary)
            diag_category_path = build_category_stack(self.last_diagnostics)
            diag_compliance_path = build_compliance_bars(
                self.transient_status,
                self._get_pn_value(),
                self._get_pmin_value(),
                self._threshold_internal_l(),
            )
            if self.air_valve_sizer.profile:
                diag_profile_path = build_profile_chart(
                    profile=self.air_valve_sizer.profile,
                    ventouses=self.air_valve_sizer.ventouses,
                    vidanges=self.air_valve_sizer.vidanges,
                    pipe_dn_mm=self.air_valve_sizer.pipe_dn_mm,
                )
        except Exception:
            # Ne pas bloquer l'export si un graphe échoue
            pass
        chart_paths.extend([diag_kpi_path, diag_category_path,
                            diag_compliance_path, diag_profile_path])

        metadata = {
            "nom_projet":       self.entry_projet.get().strip()    if hasattr(self, 'entry_projet')    else "",
            "ingenieur":        self.entry_ingenieur.get().strip() if hasattr(self, 'entry_ingenieur') else "",
            "date":             datetime.now().strftime("%d/%m/%Y"),
            "station_filepath": self.station_filepath,
            "hpt_filepath":     self.hpt_filepath,
        }

        try:
            # Récupérer le résumé du classeur si disponible
            wb_summary = None
            if self.workbook_manager.sheets:
                wb_summary = self.workbook_manager.get_summary()

            pump_summaries = []
            for p in self.pump_parsers:
                if p.parsed:
                    pump_summaries.append(p.get_summary())

            gen = WordReportGenerator()
            doc = gen.generate(
                metadata              = metadata,
                steady                = self.steady_state_status,
                transient             = self.transient_status,
                pn_label              = self.var_pn.get(),
                pn_value              = self._get_pn_value(),
                pmin_label            = self.var_pmin.get(),
                pmin_value            = self._get_pmin_value(),
                flow_unit             = self.var_flow_unit.get(),
                volume_unit           = self.var_volume_unit.get(),
                volume_threshold_disp = self._volume_threshold_display(),
                chart_png_path        = chart_path,
                workbook_summary      = wb_summary,
                pump_summaries         = pump_summaries,
                air_valve_data        = {
                    "profile":   self.air_valve_sizer.profile,
                    "ventouses": self.air_valve_sizer.ventouses,
                    "vidanges":  self.air_valve_sizer.vidanges,
                    "pipe_dn_mm": self.air_valve_sizer.pipe_dn_mm,
                } if self.air_valve_sizer.profile else None,
                diagnostics_checks    = self.last_diagnostics,
                diagnostics_summary   = self.last_diagnostics_summary,
                diag_kpi_chart_path   = diag_kpi_path,
                diag_category_chart_path = diag_category_path,
                diag_compliance_chart_path = diag_compliance_path,
                diag_profile_chart_path = diag_profile_path,
            )
            doc.save(filepath)
            messagebox.showinfo("Export réussi",
                                f"Rapport Word enregistré :\n{filepath}")
        except Exception as exc:
            messagebox.showerror("Erreur génération Word", str(exc))
        finally:
            # Nettoyage de tous les fichiers temporaires PNG
            for p in chart_paths:
                if p and os.path.exists(p):
                    try:
                        os.unlink(p)
                    except Exception:
                        pass


# =====================================================================
# 4. POINT D'ENTRÉE DU SCRIPT
# =====================================================================

if __name__ == "__main__":
    app = HammerPyApp()
    app.mainloop()
